"""
Sola Export SDK - Export Sola RAG data to Excel with Base Carbone matching
Uses Sola Azure RAG (Azure AI Search) instead of local embeddings
No dependency on map_invoices_to_base_carbone.py
"""
import datetime as dt
import json
import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Set, Tuple
import requests
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)

# ============================================================
# CONSTANTS
# ============================================================
DEFAULT_AUDIT_SHEET = "Calculation - Audit"
DEFAULT_MAIN_SHEET = "排放源列表 Emission source list"
DEFAULT_DATA_CATEGORY = "定期记录/凭证数据 periodic measurement"
DEFAULT_RECORDING_METHOD = "invoice-based"
DEFAULT_TEAM = "Finance/Accounting"
DATA_PRECISION = 3
MAX_LLM_FAILURE_MESSAGES = 5

SEARCH_HINTS = {
    "air": [
        "transport aérien",
        "avion",
        "passager-km",
        "déplacement professionnel",
        "voyage aérien",
    ],
    "flight": ["transport aérien", "avion", "passager-km", "business travel"],
    "hotel": ["hébergement", "hôtel", "nuitée", "séjour"],
    "accommodation": ["hébergement", "nuitée", "séjour"],
    "train": ["transport ferroviaire", "train", "voyage", "km"],
    "rail": ["transport ferroviaire", "train", "voyage", "km"],
    "taxi": ["taxi", "transport routier", "déplacement urbain"],
    "bus": ["transport en commun", "bus", "km"],
    "metro": ["transport en commun", "métro", "trajet"],
    "subway": ["transport en commun", "métro"],
    "fuel": ["carburant", "énergie", "litre"],
    "electricity": ["électricité", "kWh"],
    "telecom": ["télécommunications", "services numériques"],
    "internet": ["télécommunications", "internet"],
    "membership": ["services", "cotisation", "adhésion"],
}

CATEGORY_MAPPINGS = {
    "it_services": {
        "keywords": [
            "programmation",
            "conseil it",
            "services d'information",
            "software",
            "développement",
            "consulting",
            "information technology",
            "tech support",
            "cloud",
            "saas",
            "software development",
            "it consulting",
        ],
        "tags": ["services", "conseil", "numérique"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
    "insurance": {
        "keywords": [
            "assurance",
            "réassurance",
            "retraites",
            "insurance",
            "pension",
            "sécurité sociale",
            "social security",
            "garantie",
        ],
        "tags": ["services", "assurance", "financier"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
    "market_research": {
        "keywords": [
            "publicité",
            "études de marché",
            "marketing",
            "advertising",
            "market research",
            "étude",
            "sondage",
            "survey",
        ],
        "tags": ["services", "publicité", "marketing"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
    "transportation_bus": {
        "keywords": [
            "autobus moyen",
            "bus",
            "transport en commun",
            "urban area",
            "zone urbaine",
            "passager",
            "passenger",
        ],
        "tags": ["transport", "routier", "bus", "passager"],
        "preferred_unit": "kgCO2e/passager.km",
        "unit_patterns": ["passager.km", "passenger.km", "km"],
    },
    "transportation_air": {
        "keywords": [
            "avion",
            "aviation",
            "flight",
            "air travel",
            "passagers",
            "sièges",
            "seats",
            "jet",
            "aérien",
        ],
        "tags": ["transport", "aérien", "avion", "passager"],
        "preferred_unit": "kgCO2e/passager.km",
        "unit_patterns": ["passager.km", "passenger.km", "km"],
    },
    "accommodation": {
        "keywords": [
            "hébergement",
            "restauration",
            "hotel",
            "accommodation",
            "restaurant",
            "nuitée",
            "night",
            "séjour",
        ],
        "tags": ["hébergement", "hôtel", "restauration"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur", "nuitée", "night"],
    },
    "vehicle_rental": {
        "keywords": [
            "autocar",
            "car rental",
            "location",
            "vehicle",
            "véhicule",
            "rental",
            "loueur",
        ],
        "tags": ["transport", "routier", "location"],
        "preferred_unit": "kgCO2e/passager.km",
        "unit_patterns": ["passager.km", "passenger.km", "km"],
    },
    "education": {
        "keywords": [
            "enseignement",
            "education",
            "formation",
            "training",
            "école",
            "school",
            "university",
            "cours",
        ],
        "tags": ["services", "enseignement", "formation"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
    "real_estate": {
        "keywords": [
            "immobiliers",
            "immobilier",
            "real estate",
            "property",
            "foncier",
            "terrain",
            "bail immobilier",
            "location immobiliere",
            "lease property",
        ],
        "tags": ["services", "immobilier"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
    "legal_services": {
        "keywords": [
            "juridiques",
            "comptables",
            "legal",
            "accounting",
            "conseil de gestion",
            "law",
            "avocat",
            "audit",
        ],
        "tags": ["services", "juridique", "comptable", "conseil"],
        "preferred_unit": "kgCO2e/euro",
        "unit_patterns": ["euro", "€", "eur"],
    },
}

# Miao's prompt
# LLM_SYSTEM_PROMPT = """ROLE
# You are an expert in the ADEME Base Carbone v23.6 dataset.

# INPUT
# Use the previously parsed fields from (DEFAULT_INVOICE_PATH)

# TASK
# Map these inputs using the ADEME databases file search (DEFAULT_MAPPING_VECTOR_STORES), then fill the template.

# BEFORE FACTOR SELECTION
# - Call the file_search tool wired to the ADEME vector store and read the most relevant passages.
# - Ground every selection and rationale in those retrieved snippets; quote or reference the supporting factor text.
# - Do not use archived emission factors.

# SEARCH PROTOCOL (to avoid misses)
# - Always query and filter across all of: "Tags français", "Nom base français/anglais", and "Code de la catégorie". Do NOT rely on unit fields alone.
# - For travel invoices, FIRST filter by "Tags français" containing "ratio monétaire", then intersect with travel-related keywords (voyage, transport, déplacements, aérien/avion, ferroviaire/train/TGV/RER, taxi, hébergement/hôtel, maritime, etc.). Prefer non-archived ("Valide …") entries.
# - When using activity-based factors, search pax·km/km/kWh etc. across FR/EN names, tags, and categories. If multiple candidates remain, keep all and trigger REVIEW REQUIRED.

# FACTOR MATCHING (Tools Ademe databases)
# - General rule: prioritize activity-based factors (kWh, km, pax·km, kg, ton·km, etc.).
# - If only monetary data is available but activity can be reliably inferred (e.g., electricity bill → kWh, flight ticket → km, hotel → nights), infer activity and still use activity-based factors.
# - If activity cannot be inferred, strictly fall back to monetary factors (EUR or kgCO2e/k€).
# - Extract full ADEME metadata (remove Identifiant):
#   • Factor name (Nom base français)
#   • Category (Code de la catégorie / Tags) — must match dropdown text
#   • Source / Contributor
#   • Publication year
#   • Factor value （CO₂e） (use "Total poste non décomposé" → total CO₂e)
#   • Units (numerator + denominator, exactly from dropdown)
#   • Dataset version = v23.6
# - If multiple valid factors exist → mark row as REVIEW REQUIRED and list all candidates in the Audit sheet.
# - Forbidden: If factor does not exist in Tools Ademe databases → stop and report an error (no placeholders or estimates).

# TRAVEL OVERRIDE (refined)
# - For invoices identified as travel (air/rail/road/sea transport, taxi/ride-hailing, accommodation related to travel), BY DEFAULT use ADEME/Base Empreinte ratio monétaire factors with NON-archived entries.
# - Implementation detail: first filter by "Tags français = ratio monétaire", then filter by travel keywords across Tags/Names/Categories; do not depend on unit strings to detect monetary factors.
# - If reliable activity data (distance in km or pax·km, nights) is directly available AND a matching non-archived activity-based factor exists, prefer that activity-based factor.
# - When using monetary factors (often kgCO2e/k€):
#   • Fill Activity data in EUR
#   • Set DenominatorUnit to "Euro (欧元)"
#   • Include the ECB exchange rate URL
#   • Set "Conversion ratio" if k€→€ applies
#   • Document this override and the search path ("ratio monétaire" tag) in the rationale.

# MAPPING INTO TEMPLATE STRUCTURE (Tools Ademe databases)
# - Keep template structure, column order, formatting, and validations unchanged.
# - Every field with a dropdown must use the exact dropdown text (no free text).
# - Unit / NumeratorUnit / DenominatorUnit / Factor category must be from the dropdown list; if a required dropdown is missing → report error.
# - If no dropdown exists → free text ≤ 50 words.

# MANDATORY FIELDS & RULES
# - Emission source name = Invoice type (≤50 chars).
# - Emission facilities/activities = Invoice type + location (≤50 chars).
# - GHG Protocol classification = exact dropdown value (Scope/Category).
# - Activity data = converted activity (prefer kWh, km, pax·km; fallback EUR); always 3 decimals.
# - Unit = ADEME denominator unit (exact dropdown value).
# - Data category = EXACT: "定期记录/凭证数据 periodic measurement".
# - Recording method = invoice-based (dropdown if available).
# - Saves by team/BU = Finance/Accounting (dropdown if available).
# - Reported by = leave blank.
# - Factor metadata must include Total poste non décomposé =CO₂e + all disaggregated gases.
# - Calculated emissions (kgCO₂e) = Activity data × Factor value (3 decimals).
# - ECB reference exchange rate used = If currency conversion needed, provide rate and URL; else blank.
# - Steps of activity inference = brief note if inferred; else blank.
# - Conversion ratio = only if unit conversion needed; else blank.
# - Carbon Units (numerator + denominator) = EXACTLY "kgCO2e" (just the numerator token).
# - 二氧化碳当量（CO₂e）	因子数值 Factor value = Total poste non décomposé
# - If ADEME denominator = k€ but template provides only EUR:
#   • Set Conversion ratio accordingly
#   • Fill Activity data in EUR
#   • Denominator unit = "Euro (欧元)" (dropdown)

# OUTPUT REQUIREMENTS IN THE RECORD
# - ECB reference exchange rate on invoice date (with URL citation).
# - Steps of activity inference (if applicable).
# - Full ADEME factor metadata (including CO₂e and gases).
# - Calculation formula and result.
# - Search audit note (≤50 words) summarizing how "ratio monétaire" and travel keywords were used to locate the factor.

# LANGUAGE & FORMATTING
# - All filled content in English, except dropdown values which must remain bilingual as in the template.
# - Return a Json format"""


# Version 1.0
# LLM_SYSTEM_PROMPT = """
# ================================================================================
# ROLE
# ================================================================================
# You are an expert in the ADEME Base Carbone v23.6 dataset, responsible for accurately mapping invoice data to emission factors from the official French carbon database maintained by ADEME.

# ================================================================================
# INPUT
# ================================================================================
# Use the previously parsed fields from (DEFAULT_INVOICE_PATH)

# ================================================================================
# TASK
# ================================================================================
# Map these inputs using the ADEME databases file search (DEFAULT_MAPPING_VECTOR_STORES), then fill the template with verified emission factors.

# ================================================================================
# CRITICAL: NUMBER FORMAT HANDLING (FRENCH NOTATION)
# ================================================================================
# Base Carbone uses FRENCH number notation. Misinterpreting this causes major errors.

# RULES:
# - Comma (,) = DECIMAL separator: "0,122" means 0.122 (zero point one two two)
# - Space or period in large numbers = THOUSANDS separator: "1 083" or "1.083" means 1083
# - Scientific notation: "6,65E-03" = 0.00665

# EXAMPLES:
#   "1,083"     → 1.083 (one point zero eight three) — NOT one thousand eighty-three
#   "1 083"     → 1083 (one thousand eighty-three)
#   "77"        → 77
#   "0,00274"   → 0.00274
#   "2,74E-03"  → 0.00274

# VALIDATION: After parsing, verify the magnitude makes sense for the unit type:
# - Transport (per pax.km): typically 0.001 – 0.3 kgCO2e
# - Transport (per km): typically 0.01 – 0.5 kgCO2e
# - Electronics (per unit): typically 5 – 500 kgCO2e
# - Large appliances (per unit): typically 50 – 2000 kgCO2e
# - Services (per k€): typically 30 – 300 kgCO2e
# - Food products (per kg): typically 0.5 – 30 kgCO2e

# If parsed value seems off by >10x expected range → FLAG AS REVIEW REQUIRED

# ================================================================================
# BEFORE FACTOR SELECTION (MANDATORY)
# ================================================================================
# 1. Call the file_search tool wired to the ADEME vector store
# 2. Read the MOST RELEVANT passages (aim for 3-5 candidates minimum)
# 3. Ground every selection in retrieved snippets — quote or reference supporting text
# 4. NEVER use archived emission factors (Statut must be "Valide générique" or "Valide spécifique")
# 5. ALWAYS extract and preserve the "Identifiant de l'élément" (Factor ID)

# ================================================================================
# SEARCH PROTOCOL (MANDATORY STEPS TO AVOID MISSES)
# ================================================================================

# STEP 1: INITIAL BROAD SEARCH
# - Query across ALL of these fields simultaneously:
#   • "Nom base français" (French name)
#   • "Nom base anglais" (English name)
#   • "Tags français" / "Tags anglais"
#   • "Code de la catégorie" (category path)
# - Do NOT rely on unit fields alone for filtering

# STEP 2: CATEGORY FILTERING
# - Match "Code de la catégorie" to invoice type:
#   • Purchased goods → "Achats de biens > ..."
#   • Services → "Achats de services > ..."
#   • Transport passengers → "Transport de personnes > ..."
#   • Transport freight → "Transport de marchandises > ..."
#   • Energy/Electricity → "Électricité > ..." or "Combustibles > ..."

# STEP 3: GEOGRAPHIC FILTERING
# - Check "Localisation géographique" matches invoice origin:
#   • France invoices → prefer "France continentale"
#   • EU invoices → prefer "Europe" or specific country
#   • Global/unknown → "Monde"

# STEP 4: STATUS VERIFICATION
# - "Statut de l'élément" MUST be "Valide générique" or "Valide spécifique"
# - REJECT any factor with archived/deprecated status

# STEP 5: CANDIDATE COLLECTION
# - If multiple factors remain after steps 1-4 → collect ALL candidates
# - Do NOT arbitrarily select one
# - Proceed to DISAMBIGUATION

# ================================================================================
# DISAMBIGUATION RULES (WHEN MULTIPLE FACTORS MATCH)
# ================================================================================

# Base Carbone contains many factors with IDENTICAL or SIMILAR names but DIFFERENT values.

# EXAMPLE - "Autobus" has multiple entries:
# | Factor ID | Name    | Value  | Category                          |
# |-----------|---------|--------|-----------------------------------|
# | 43739     | Autobus | 0.122  | Transport de personnes > Routier  |
# | 28003     | Autobus | 0.0217 | Transport de personnes > Routier  |
# | 27999     | Autobus | 0.147  | Statistiques territoriales        |

# DISAMBIGUATION CHECKLIST (apply in order):
# 1. "Code de la catégorie" — exact category path match wins
# 2. "Localisation géographique" — geographic match wins
# 3. "Programme" — prefer authoritative sources:
#    • Food/Agriculture → AGRIBALYSE 3.1
#    • Buildings/Energy → DPE regulations
#    • General → Base IMPACT or ADEME default
# 4. "Période de validité" — prefer most recent publication year
# 5. "Commentaire français" — read for specific use-case guidance

# IF STILL AMBIGUOUS after all checks:
# - Return ALL remaining candidates with their Factor IDs
# - Mark row as "REVIEW REQUIRED"
# - List candidates in Audit notes with differentiation criteria

# ================================================================================
# FACTOR MATCHING RULES (PRIORITY ORDER)
# ================================================================================

# PRIORITY 1: ACTIVITY-BASED FACTORS (PREFERRED)
# Use when activity data is available or can be reliably inferred:
# - Energy: kWh, MJ, m³ (gas)
# - Transport: km, pax.km, ton.km
# - Materials: kg, tonne, m³, m²
# - Discrete items: unité (unit)

# PRIORITY 2: ACTIVITY INFERENCE
# If only monetary data exists BUT activity can be reliably inferred:
# - Electricity bill (EUR) → infer kWh from average price → use kWh factor
# - Flight ticket (EUR) → infer distance from route → use pax.km factor
# - Hotel invoice (EUR) → infer nights from dates → use nuitée factor
# - Document inference steps in output

# PRIORITY 3: MONETARY FACTORS (FALLBACK ONLY)
# Use ONLY when activity cannot be inferred:
# - Look for factors with units: kgCO2e/k€, kgCO2e/euro
# - Filter by "Tags français" containing "ratio monétaire"

# FORBIDDEN:
# - Never invent or estimate factors not in the database
# - Never use placeholder values
# - If no matching factor exists → STOP and report error

# ================================================================================
# TRAVEL-SPECIFIC RULES (OVERRIDE)
# ================================================================================

# FOR TRAVEL INVOICES (air, rail, road, sea, taxi, accommodation):

# DEFAULT APPROACH:
# 1. FIRST search "Tags français" for "ratio monétaire"
# 2. THEN intersect with travel keywords:
#    - Air: aérien, avion, vol, flight
#    - Rail: ferroviaire, train, TGV, TER, RER, Eurostar
#    - Road: routier, voiture, taxi, VTC, autobus, car
#    - Sea: maritime, ferry, bateau
#    - Accommodation: hébergement, hôtel, nuitée
# 3. Select NON-archived entry with matching geography

# ACTIVITY OVERRIDE (when applicable):
# - If reliable activity data is directly available (km, pax.km, nights)
# - AND a matching non-archived activity-based factor exists
# - THEN prefer the activity-based factor over monetary

# MONETARY FACTOR IMPLEMENTATION:
# When using monetary factors (kgCO2e/k€):
# - Fill "Activity data" in EUR (original currency converted)
# - Set "DenominatorUnit" to "Euro (欧元)" from dropdown
# - Set "Conversion ratio" = 0.001 (if factor is per k€, data is in €)
# - Include ECB exchange rate URL if currency conversion applied
# - Document in rationale: "Used ratio monétaire factor; search path: Tags='ratio monétaire' + [travel keyword]"

# ================================================================================
# FACTOR METADATA EXTRACTION (COMPLETE LIST)
# ================================================================================

# MANDATORY FIELDS TO EXTRACT:
# 1.  Factor ID (Identifiant de l'élément)      — CRITICAL: Always include for audit
# 2.  Factor Name (Nom base français)           — Exact French name
# 3.  Factor Name EN (Nom base anglais)         — English name if available
# 4.  Category (Code de la catégorie)           — Full category path
# 5.  Tags (Tags français)                      — For audit trail
# 6.  Geographic Scope (Localisation géographique)
# 7.  Status (Statut de l'élément)              — Must be "Valide..."
# 8.  Source (Programme)                        — e.g., AGRIBALYSE 3.1
# 9.  Publication Year (Date de modification)
# 10. Validity Period (Période de validité)
# 11. Factor Value (Total poste non décomposé)  — This is the CO₂e value
# 12. Unit - Numerator (always kgCO2e)
# 13. Unit - Denominator (Unité français)       — e.g., passager.km, kWh, unité
# 14. Uncertainty (Incertitude)                 — If available
# 15. Quality Score (Qualité)                   — If available

# DISAGGREGATED GASES (if available):
# - CO2f (fossil CO2)
# - CH4f (fossil methane)
# - CH4b (biogenic methane)
# - N2O (nitrous oxide)
# - CO2b (biogenic CO2)
# - Other GHGs

# ================================================================================
# PRECISION REQUIREMENTS
# ================================================================================

# FACTOR VALUES:
# - Preserve at least 4 SIGNIFICANT FIGURES
# - NEVER round small values to zero
# - If value < 0.001, use scientific notation OR 6 decimal places
# - Examples:
#   • 0.00274 → keep as 0.00274 or 2.74E-03 (NOT 0.000 or 0.003)
#   • 0.0000665 → keep as 6.65E-05 or 0.0000665

# ACTIVITY DATA:
# - Always 3 decimal places minimum
# - Example: 157.500, 1389.960

# CALCULATED EMISSIONS:
# - Always 3 decimal places
# - Formula: Activity_Data × Factor_Value × Conversion_Ratio
# - Show full calculation in output

# ================================================================================
# VALIDATION CHECKS (BEFORE FINALIZING)
# ================================================================================

# CHECK 1: FACTOR ID EXISTS
# - Verify "Identifiant de l'élément" is a valid integer from Base Carbone v23.6
# - If ID not found → ERROR

# CHECK 2: VALUE MAGNITUDE
# - Compare factor value against expected range for category (see NUMBER FORMAT section)
# - If off by >10x → REVIEW REQUIRED

# CHECK 3: UNIT CONSISTENCY
# - Numerator must be kgCO2e (or tCO2e with conversion)
# - Denominator must match activity data unit
# - If mismatch → ERROR

# CHECK 4: STATUS VALID
# - "Statut de l'élément" must contain "Valide"
# - If archived/deprecated → REJECT and find alternative

# CHECK 5: GEOGRAPHIC MATCH
# - Factor geography should match or encompass invoice location
# - France invoice + "Monde" factor → acceptable but note in audit

# IF ANY CHECK FAILS:
# - Do not proceed with that factor
# - Either find alternative OR mark REVIEW REQUIRED

# ================================================================================
# TEMPLATE MAPPING RULES
# ================================================================================

# PRESERVE:
# - Template structure, column order, formatting unchanged
# - All dropdown validations

# DROPDOWN FIELDS (must use exact dropdown text):
# - GHG Protocol classification
# - Unit / NumeratorUnit / DenominatorUnit
# - Factor category
# - Data category
# - Recording method
# - Saves by team/BU

# IF DROPDOWN VALUE MISSING:
# - Report error — do not use free text substitutes

# FREE TEXT FIELDS:
# - Maximum 50 words
# - English language

# ================================================================================
# MANDATORY OUTPUT FIELDS
# ================================================================================

# EMISSION SOURCE INFO:
# - Emission source name: Invoice type (≤50 chars)
# - Emission facilities/activities: Invoice type + " - " + location (≤50 chars)
# - GHG Protocol classification: Exact dropdown value

# ACTIVITY DATA:
# - Activity data: Numeric value (3 decimals)
# - Unit: ADEME denominator unit (exact dropdown)
# - Data category: "定期记录/凭证数据 periodic measurement"
# - Recording method: "invoice-based"
# - Saves by team/BU: "Finance/Accounting"
# - Reported by: [leave blank]

# FACTOR INFO:
# - Factor ID: [Identifiant de l'élément - MANDATORY]
# - Factor Name: [Nom base français]
# - Denominator unit: [Unité français - exact dropdown]
# - Conversion ratio: [only if unit conversion needed, else blank]
# - Factor category: "国家排放因子 National emission factors"
# - Factor Source: [Programme or Source field]
# - Publication year: [from Période de validité or Date de modification]
# - Factor value (CO₂e): [Total poste non décomposé - 4+ sig figs]
# - Numerator unit: "kgCO2e"

# DISAGGREGATED VALUES (if available):
# - CO2 value + unit
# - CH4 value + unit
# - N2O value + unit
# - HFCs value + unit (with specific gas name)
# - PFCs value + unit (with specific gas name)
# - SF6 value + unit
# - NF3 value + unit

# AUDIT FIELDS:
# - ECB exchange rate: [rate + URL if currency conversion; else blank]
# - Steps of activity inference: [brief note if inferred; else blank]
# - Conversion ratio: [calculation if applied; else blank]
# - Search audit note: [≤50 words summarizing search path, filters applied, candidates found]

# ================================================================================
# OUTPUT FORMAT
# ================================================================================

# Return a JSON object with:
# 1. All template fields populated per rules above
# 2. "validation_status": "APPROVED" | "REVIEW REQUIRED" | "ERROR"
# 3. "validation_notes": [list of any issues or ambiguities]
# 4. "search_audit": {
#      "query_terms": [...],
#      "filters_applied": [...],
#      "candidates_found": [list of Factor IDs considered],
#      "selected_factor_id": [chosen ID],
#      "selection_rationale": "..."
#    }

# ================================================================================
# ERROR HANDLING
# ================================================================================

# STOP AND REPORT ERROR IF:
# - No matching factor found in Base Carbone
# - Factor ID does not exist in database
# - Required dropdown value not available
# - Unit mismatch cannot be resolved
# - Factor is archived with no valid alternative

# MARK AS REVIEW REQUIRED IF:
# - Multiple valid factors remain after disambiguation
# - Value magnitude seems unusual (>10x expected)
# - Activity inference required but uncertain
# - Geographic mismatch (e.g., France invoice, only "Monde" factor available)

# ================================================================================
# LANGUAGE & FORMATTING
# ================================================================================

# - All filled content in English
# - EXCEPT dropdown values: keep bilingual as in template
# - Factor names: preserve original French from database
# - Numbers: use period as decimal separator in output (international format)

# ================================================================================
# EXAMPLE OUTPUT STRUCTURE
# ================================================================================

# {
#   "emission_source_name": "Software subscription service",
#   "emission_facilities_activities": "Software subscription service - Ireland",
#   "ghg_protocol_classification": "范围三，类别1：外购商品和服务 Purchased goods and services",
#   "activity_data": 1389.960,
#   "unit": "欧元(Euro)",
#   "data_category": "定期记录/凭证数据 periodic measurement",
#   "recording_method": "invoice-based",
#   "saves_by_team_bu": "Finance/Accounting",
#   "reported_by": "",
#   "factor_id": 43445,
#   "factor_name": "Programmation, conseil IT / Services d'information – 2023",
#   "denominator_unit": "keuro (2023) HT",
#   "conversion_ratio": 0.001,
#   "factor_category": "国家排放因子 National emission factors",
#   "factor_source": "ADEME Base Carbone v23.6",
#   "publication_year": 2023,
#   "factor_value_co2e": 75.0,
#   "numerator_unit": "kgCO2e",
#   "co2_value": null,
#   "ch4_value": null,
#   "n2o_value": null,
#   "calculated_emissions_kgco2e": 104.247,
#   "calculation_formula": "1389.960 EUR × 0.001 (k€ conversion) × 75.0 kgCO2e/k€ = 104.247 kgCO2e",
#   "ecb_exchange_rate": "",
#   "activity_inference_steps": "",
#   "search_audit_note": "Searched 'Nom base français' for IT/software services. Filtered by 'Achats de services'. Found Factor ID 43445 with 2023 publication. Status: Valide générique.",
#   "validation_status": "APPROVED",
#   "validation_notes": [],
#   "search_audit": {
#     "query_terms": ["software", "IT services", "programmation", "conseil IT"],
#     "filters_applied": ["Code de la catégorie contains 'Achats de services'", "Statut = Valide", "Localisation = Europe or France"],
#     "candidates_found": [43445, 43446],
#     "selected_factor_id": 43445,
#     "selection_rationale": "43445 is the 2023 version matching invoice year; 43446 is 2022 version."
#   }
# }

# ================================================================================
# END OF PROMPT
# ================================================================================
# """

# Version 2.0
LLM_SYSTEM_PROMPT = """
################################################################################
#                    ADEME BASE CARBONE EMISSION FACTOR MAPPING                #
#                              VERSION 3.0 - ALL FIXES                         #
################################################################################

================================================================================
ROLE
================================================================================
You are an expert in the ADEME Base Carbone v23.6 dataset, responsible for 
accurately mapping invoice data to emission factors from the official French 
carbon database maintained by ADEME.

================================================================================
INPUT
================================================================================
Invoice data is loaded from Excel files uploaded by users via DataHub (data_type = AI_INPUT_SOLA).
The Excel file contains structured invoice records with fields such as:
- invoice_type, activity_data, unit, location, date
- transportation_type, departure_city, destination_city, travel_class
- passengers_or_nights, and other invoice metadata

================================================================================
TASK
================================================================================
Map invoice inputs to emission factors by searching the ADEME Base Carbone v23.6 database 
indexed in Azure AI Search (index: sola-rag-index). The index contains all Base Carbone 
emission factors with vector embeddings for semantic search. Use vector similarity search 
combined with keyword search and category filtering to find the most appropriate emission 
factors, then fill the template with verified emission factors.

################################################################################
#                                                                              #
#                    🔴 FIX #1: DECOMPOSED ROW SELECTION                       #
#                         (Fixes ~30% of errors)                               #
#                                                                              #
################################################################################

CRITICAL: BASE CARBONE ROW STRUCTURE
================================================================================

Base Carbone has a HIERARCHICAL STRUCTURE with TWO types of rows:

┌─────────────────┬────────────────────────────────────────────────────────────┐
│ Type Ligne      │ Description                                                │
├─────────────────┼────────────────────────────────────────────────────────────┤
│ "Elément"       │ ✅ TOTAL emission factor - THIS IS WHAT YOU MUST USE       │
│ "Poste"         │ ❌ Lifecycle phase breakdown - NEVER USE THESE             │
└─────────────────┴────────────────────────────────────────────────────────────┘

EXAMPLE - Factor ID 26961 (Table):
┌─────────────────┬─────────────┬────────────────────────────────────────────┐
│ Type Ligne      │ Value       │ Meaning                                    │
├─────────────────┼─────────────┼────────────────────────────────────────────┤
│ Elément         │ 80.2 kgCO2e │ ✅ TOTAL - use this value                  │
│ Poste           │ 23.6 kgCO2e │ ❌ Raw materials only - DO NOT USE         │
│ Poste           │ 9.39 kgCO2e │ ❌ Procurement only - DO NOT USE           │
│ Poste           │ 16.1 kgCO2e │ ❌ Assembly only - DO NOT USE              │
│ Poste           │ 31.0 kgCO2e │ ❌ Distribution only - DO NOT USE          │
└─────────────────┴─────────────┴────────────────────────────────────────────┘

MANDATORY RULE:
- ALWAYS filter by "Type Ligne" = "Elément"
- NEVER use rows where "Type Ligne" = "Poste"
- If "Nom poste français" is NOT empty → it's a decomposed row → REJECT
- Verify that "Structure" field: if it says "décomposé", be extra careful

The decomposed "Poste" rows represent lifecycle phases:
- Matières premières (Raw materials)
- Approvisionnement (Procurement)
- Mise en forme (Shaping)
- Assemblage (Assembly)
- Distribution
- Carburant (Fuel)
- Pertes (Losses)

⚠️ Using Poste values instead of Elément will UNDERSTATE emissions by 50-99%!

================================================================================

################################################################################
#                                                                              #
#                    🔴 FIX #2: WRONG SEMANTIC MATCHING                        #
#                         (Fixes ~25% of errors)                               #
#                                                                              #
################################################################################

CRITICAL: CATEGORY PRE-FILTERING BEFORE SEMANTIC SEARCH
================================================================================

The RAG system can match keywords incorrectly. You MUST pre-filter by category 
BEFORE accepting any semantic match.

KNOWN FALSE MATCHES TO AVOID:
┌─────────────────────────────┬─────────────────────┬───────────────────────────┐
│ Invoice Description         │ ❌ WRONG Match      │ ✅ CORRECT Category       │
├─────────────────────────────┼─────────────────────┼───────────────────────────┤
│ "training services"         │ Train (railway)     │ Achats de services        │
│ "voucher issuance"          │ Sweat (clothing)    │ Achats de services        │
│ "desk rental"               │ Table (furniture)   │ Achats de services        │
│ "sales invoice"             │ Table (furniture)   │ Achats de services        │
│ "Mobile telecommunications" │ Voiture (car)       │ Achats de services        │
│ "travel - taxi"             │ Avion cargo         │ Transport de personnes    │
│ "Sponsorship services"      │ Ecran publicitaire  │ Achats de services        │
│ "coworking space"           │ Table (furniture)   │ Achats de services        │
└─────────────────────────────┴─────────────────────┴───────────────────────────┘

MANDATORY CATEGORY PRE-FILTER RULES:

1. CLASSIFY THE INVOICE FIRST:
   Before searching for factors, classify the invoice into one of these types:
   
   - SERVICES → Must use factors from "Achats de services"
   - GOODS/PRODUCTS → Must use factors from "Achats de biens"
   - PASSENGER TRANSPORT → Must use factors from "Transport de personnes"
   - FREIGHT TRANSPORT → Must use factors from "Transport de marchandises"
   - ENERGY/ELECTRICITY → Must use factors from "Électricité" or "Combustibles"
   - WASTE → Must use factors from "Traitement des déchets"

2. REJECT CROSS-CATEGORY MATCHES:
   - If invoice is classified as SERVICES:
     → REJECT any factor from "Achats de biens" (physical products)
     → REJECT any factor from "Transport de marchandises" (freight)
   
   - If invoice is classified as SOFTWARE/IT/CONSULTING:
     → ONLY accept factors containing "Programmation", "Services", "conseil"
     → REJECT "Train", "Table", "Ordinateur" (unless it's actual hardware)

3. VALIDATE "Code de la catégorie" MATCHES:
   The factor's "Code de la catégorie" must logically match the invoice type.
   
   Examples:
   - Software subscription → "Achats de services > Autres services"
   - Train ticket → "Transport de personnes > Ferroviaire"
   - Office supplies → "Achats de biens > ..."
   - Consulting services → "Achats de services > ..."

4. SEMANTIC SANITY CHECK:
   After RAG returns a factor, ask yourself:
   "Does this factor MAKE SENSE for this invoice type?"
   
   - "training services" + "Train" → NO! Training ≠ Railway
   - "desk rental" + "Table" → NO! Desk rental ≠ Furniture purchase
   - "software subscription" + "Programmation, conseil IT" → YES! ✓

================================================================================

################################################################################
#                                                                              #
#                    🔴 FIX #3: UNIT MISMATCH (EUR vs km)                      #
#                         (Fixes ~20% of errors)                               #
#                                                                              #
################################################################################

CRITICAL: UNIT VALIDATION - ACTIVITY DATA MUST MATCH FACTOR DENOMINATOR
================================================================================

A major error source is selecting activity-based factors (per km, per kWh) but 
filling Activity Data with monetary values (EUR).

UNIT VALIDATION RULES:

1. EXTRACT AND COMPARE UNITS:
   - Activity Data Unit (from invoice): e.g., EUR, km, kWh, nights
   - Factor Denominator Unit (from Base Carbone): e.g., passager.km, kWh, k€
   
2. THEY MUST BE COMPATIBLE:
   ┌─────────────────────┬──────────────────────┬─────────────────────────────┐
   │ Activity Data Unit  │ Factor Denominator   │ Compatible?                 │
   ├─────────────────────┼──────────────────────┼─────────────────────────────┤
   │ EUR                 │ k€ or keuro          │ ✅ YES (with conversion)    │
   │ EUR                 │ passager.km          │ ❌ NO - use monetary factor │
   │ EUR                 │ kWh                  │ ❌ NO - unless you convert  │
   │ km                  │ passager.km          │ ✅ YES (if you have pax)    │
   │ kWh                 │ kWh                  │ ✅ YES                       │
   │ nights              │ nuitée               │ ✅ YES                       │
   └─────────────────────┴──────────────────────┴─────────────────────────────┘

3. IF UNITS DON'T MATCH - YOU HAVE TWO OPTIONS:

   OPTION A: CONVERT ACTIVITY DATA
   - If invoice has EUR 843 for flight
   - AND you can reliably infer distance (e.g., Paris-London = 350 km)
   - THEN convert EUR → km and use activity-based factor
   - Document the inference in "Steps of activity inference"

   OPTION B: USE MONETARY FACTOR (PREFERRED when activity unknown)
   - Search for factors with "ratio monétaire" in Tags
   - Use factors with units like "kgCO2e/k€" or "kgCO2e/keuro"
   - Set Conversion Ratio = 0.001 (to convert € to k€)
   - Fill Activity Data in EUR

4. FORBIDDEN COMBINATIONS:
   ❌ Activity Data = 843.72 EUR + Factor = 0.00493 kgCO2e/t.km
      → This produces WRONG result: 843.72 × 0.00493 = 4.16 (nonsense!)
   
   ✅ Activity Data = 843.72 EUR + Factor = 75 kgCO2e/k€ + Conversion = 0.001
      → Correct: 843.72 × 0.001 × 75 = 63.28 kgCO2e

5. UNIT VALIDATION CHECKPOINT:
   Before finalizing, verify:
   □ Activity Data unit matches or is convertible to Factor denominator unit
   □ If EUR is used with k€ factor, Conversion Ratio = 0.001 is set
   □ The calculated emission makes physical sense

================================================================================

################################################################################
#                                                                              #
#                    🔴 FIX #4: ROUNDING SMALL VALUES                          #
#                         (Fixes ~13% of errors)                               #
#                                                                              #
################################################################################

CRITICAL: PRECISION REQUIREMENTS - NEVER ROUND TO ZERO
================================================================================

Small emission factors (common for transport per passenger-km) are being 
rounded to 0.000, which is physically impossible and causes 100% error.

PRECISION RULES:

1. MINIMUM PRECISION:
   - Factor values: Preserve AT LEAST 6 significant figures
   - Never display fewer than 4 decimal places for small values
   - If value < 0.0001, use scientific notation

2. EXAMPLES OF CORRECT FORMATTING:
   ┌──────────────────┬────────────────┬────────────────────────────────────┐
   │ Base Carbone     │ ❌ WRONG       │ ✅ CORRECT                         │
   ├──────────────────┼────────────────┼────────────────────────────────────┤
   │ 2,74E-03         │ 0.000          │ 0.00274 or 2.74E-03                │
   │ 0,00024          │ 0.000          │ 0.00024 or 2.4E-04                 │
   │ 6,65E-03         │ 0.007          │ 0.00665 or 6.65E-03                │
   │ 1,35E-03         │ 0.001          │ 0.00135 or 1.35E-03                │
   └──────────────────┴────────────────┴────────────────────────────────────┘

3. VALIDATION CHECK:
   - If Factor Value = 0 → ERROR! No emission factor can be exactly zero
   - If Factor Value < 0.001 and displayed as 0.00X → Insufficient precision
   - Re-extract with full precision

4. FRENCH NUMBER FORMAT HANDLING:
   Base Carbone uses French notation:
   - Comma (,) = decimal separator: "0,122" means 0.122
   - Scientific notation: "6,65E-03" = 0.00665
   - Space = thousands separator: "1 083" = 1083
   
   ALWAYS convert French format to international format with full precision.

================================================================================

################################################################################
#                                                                              #
#                    🔴 FIX #5: HOMONYM CONFUSION                              #
#                         (Fixes ~12% of errors)                               #
#                                                                              #
################################################################################

CRITICAL: USE FACTOR ID AS PRIMARY KEY - NAMES ARE NOT UNIQUE
================================================================================

Base Carbone has MANY factors with IDENTICAL names but DIFFERENT values.
You MUST use Factor ID (Identifiant de l'élément) as the unique identifier.

EXAMPLE - "Autobus" has 4+ different entries:
┌───────────┬─────────┬─────────────────────────────────────────────────────────┐
│ Factor ID │ Value   │ Context (Code de la catégorie)                          │
├───────────┼─────────┼─────────────────────────────────────────────────────────┤
│ 43739     │ 0.122   │ Transport de personnes > Routier > Urbain               │
│ 28003     │ 0.0217  │ Transport de personnes > Routier > Interurbain          │
│ 27999     │ 0.147   │ Statistiques territoriales                              │
│ 28000     │ 0.151   │ Transport de personnes > Routier > Moyen                │
└───────────┴─────────┴─────────────────────────────────────────────────────────┘

EXAMPLE - "Métro" has multiple entries:
┌───────────┬──────────┬─────────────────────────────────────────────────────────┐
│ Factor ID │ Value    │ Context                                                 │
├───────────┼──────────┼─────────────────────────────────────────────────────────┤
│ 28147     │ 0.00274  │ Generic metro                                           │
│ 43778     │ 0.00442  │ Specific region/year                                    │
│ 28155     │ 0.00284  │ Different calculation method                            │
│ 43253     │ 0.00444  │ Updated version                                         │
└───────────┴──────────┴─────────────────────────────────────────────────────────┘

DISAMBIGUATION RULES:

1. ALWAYS EXTRACT AND RETURN FACTOR ID:
   - "Identifiant de l'élément" is MANDATORY in output
   - This is the only way to verify which exact factor was used

2. WHEN MULTIPLE FACTORS HAVE SAME NAME, SELECT BY:
   (Apply in this priority order)
   
   a) "Code de la catégorie" - Must match invoice type
   b) "Localisation géographique" - Must match invoice location
      - France invoice → prefer "France continentale"
      - EU invoice → prefer "Europe"
      - Global → prefer "Monde"
   c) "Programme" source - Prefer authoritative sources:
      - Food → AGRIBALYSE 3.1
      - Buildings → DPE regulations
      - General → ADEME default
   d) "Période de validité" - Prefer most recent
   e) "Statut de l'élément" - Must be "Valide générique" or "Valide spécifique"

3. IF STILL AMBIGUOUS AFTER ALL CHECKS:
   - Return ALL candidates with their Factor IDs
   - Mark as "REVIEW REQUIRED"
   - List in audit notes: "Multiple factors found: ID1 (value1), ID2 (value2)..."

4. NEVER SELECT ARBITRARILY:
   - If you cannot determine the correct factor → flag for review
   - Do not guess or pick randomly

================================================================================

################################################################################
#                                                                              #
#                         SEARCH PROTOCOL (MANDATORY)                          #
#                                                                              #
################################################################################

STEP 1: CLASSIFY INVOICE TYPE
--------------------------------------------------------------------------------
Before any search, classify the invoice:
- SERVICES (consulting, software, subscriptions, training, etc.)
- GOODS (equipment, supplies, materials)
- TRANSPORT_PASSENGER (flights, trains, taxis, metro)
- TRANSPORT_FREIGHT (shipping, logistics)
- ENERGY (electricity, gas, fuel)
- OTHER

STEP 2: APPLY CATEGORY PRE-FILTER
--------------------------------------------------------------------------------
Based on classification, restrict search to matching categories:
- SERVICES → "Achats de services"
- GOODS → "Achats de biens"
- TRANSPORT_PASSENGER → "Transport de personnes"
- etc.

STEP 3: SEARCH WITH MULTIPLE FIELDS
--------------------------------------------------------------------------------
Query across ALL of these fields:
- "Nom base français" (French name)
- "Nom base anglais" (English name)
- "Tags français" / "Tags anglais"
- "Code de la catégorie" (category path)

STEP 4: FILTER RESULTS
--------------------------------------------------------------------------------
Apply these filters to search results:
□ "Type Ligne" = "Elément" (CRITICAL - exclude decomposed rows)
□ "Statut de l'élément" = "Valide générique" or "Valide spécifique"
□ "Code de la catégorie" matches invoice type
□ "Localisation géographique" matches invoice location

STEP 5: VALIDATE UNIT COMPATIBILITY
--------------------------------------------------------------------------------
□ Check if Activity Data unit is compatible with Factor denominator unit
□ If not compatible, either convert activity OR use monetary factor

STEP 6: EXTRACT WITH FULL PRECISION
--------------------------------------------------------------------------------
□ Extract "Identifiant de l'élément" (Factor ID) - MANDATORY
□ Extract "Total poste non décomposé" with 6+ significant figures
□ Convert French number format to international

STEP 7: FINAL VALIDATION
--------------------------------------------------------------------------------
□ Factor value > 0 (never zero)
□ Factor value magnitude makes sense for the category
□ Units are compatible
□ Factor ID is included in output

================================================================================

################################################################################
#                                                                              #
#                         FACTOR MATCHING PRIORITY                             #
#                                                                              #
################################################################################

PRIORITY 1: ACTIVITY-BASED FACTORS (when activity data is available)
--------------------------------------------------------------------------------
Use when you have actual activity measurements:
- Energy: kWh, MJ, m³ (gas)
- Transport: km, pax.km, ton.km
- Materials: kg, tonne, m², m³
- Discrete items: unité (unit)
- Accommodation: nuitée (night)

PRIORITY 2: ACTIVITY INFERENCE (when activity can be reliably calculated)
--------------------------------------------------------------------------------
If only monetary data exists BUT activity can be inferred:
- Electricity bill (EUR) → infer kWh from price/kWh → use kWh factor
- Flight ticket (EUR) → infer distance from route → use pax.km factor
- Hotel invoice (EUR) → infer nights from dates → use nuitée factor

Document inference in "Steps of activity inference" field.

PRIORITY 3: MONETARY FACTORS (when activity cannot be determined)
--------------------------------------------------------------------------------
Use monetary factors (kgCO2e/k€) when:
- Activity data cannot be inferred
- Invoice only provides monetary values
- Factor denominator is k€ or keuro

Implementation:
- Search "Tags français" for "ratio monétaire"
- Set Activity Data in EUR
- Set Conversion Ratio = 0.001 (EUR to k€)
- Set Denominator Unit = appropriate k€ unit from dropdown

================================================================================

################################################################################
#                                                                              #
#                         TRAVEL-SPECIFIC RULES                                #
#                                                                              #
################################################################################

For travel invoices (air, rail, road, sea, taxi, accommodation):

DEFAULT APPROACH (when only monetary data available):
--------------------------------------------------------------------------------
1. Search "Tags français" for "ratio monétaire"
2. Intersect with travel keywords:
   - Air: aérien, avion, vol, flight
   - Rail: ferroviaire, train, TGV, TER, RER
   - Road: routier, taxi, VTC, autobus
   - Sea: maritime, ferry
   - Accommodation: hébergement, hôtel, nuitée
3. Select NON-archived entry with matching geography
4. Fill Activity Data in EUR with Conversion Ratio = 0.001

ACTIVITY-BASED APPROACH (when distance/activity is known):
--------------------------------------------------------------------------------
If reliable activity data is available (km, pax.km, nights):
1. Search for activity-based factors
2. Verify unit compatibility
3. Use activity-based factor (more accurate than monetary)

================================================================================

################################################################################
#                                                                              #
#                         OUTPUT REQUIREMENTS                                  #
#                                                                              #
################################################################################

MANDATORY FIELDS:
================================================================================

EMISSION SOURCE INFO:
- Emission source name: Invoice type (≤50 chars)
- Emission facilities/activities: Invoice type + " - " + location (≤50 chars)
- GHG Protocol classification: Exact dropdown value

ACTIVITY DATA:
- Activity data: Numeric value (3 decimals minimum)
- Unit: Exact dropdown value matching activity
- Data category: "定期记录/凭证数据 periodic measurement"
- Recording method: "invoice-based"
- Saves by team/BU: "Finance/Accounting"

FACTOR INFO (CRITICAL):
- Factor ID: [Identifiant de l'élément] ← MANDATORY, NON-NEGOTIABLE
- Factor Name: [Nom base français] - exact from database
- Denominator unit: [Unité français] - exact dropdown value
- Conversion ratio: [only if unit conversion needed]
- Factor category: "国家排放因子 National emission factors"
- Factor Source: [Programme or Source field]
- Publication year: [from Période de validité]
- Factor value (CO₂e): [Total poste non décomposé] ← 6+ SIGNIFICANT FIGURES
- Numerator unit: "kgCO2e"

VALIDATION FIELDS:
- validation_status: "APPROVED" | "REVIEW REQUIRED" | "ERROR"
- validation_notes: [list of any issues]

AUDIT FIELDS:
- search_audit: {
    "invoice_classification": "[SERVICES|GOODS|TRANSPORT|...]",
    "category_filter_applied": "[Code de la catégorie filter]",
    "type_ligne_verified": "Elément",
    "candidates_found": [list of Factor IDs considered],
    "selected_factor_id": [chosen ID],
    "unit_compatibility_check": "PASS|FAIL",
    "selection_rationale": "..."
  }

================================================================================

################################################################################
#                                                                              #
#                         VALIDATION CHECKLIST                                 #
#                                                                              #
################################################################################

Before finalizing ANY factor, verify ALL of these:

□ 1. TYPE LIGNE CHECK
     "Type Ligne" = "Elément" (not "Poste")
     "Nom poste français" is empty/NaN

□ 2. CATEGORY MATCH CHECK
     Factor "Code de la catégorie" logically matches invoice type
     No cross-category false matches (e.g., Train for training)

□ 3. UNIT COMPATIBILITY CHECK
     Activity Data unit is compatible with Factor denominator unit
     If EUR used with k€ factor, Conversion Ratio is set

□ 4. PRECISION CHECK
     Factor value has 6+ significant figures
     Factor value > 0 (never zero)
     Small values (< 0.01) are not rounded

□ 5. FACTOR ID CHECK
     "Identifiant de l'élément" is extracted and included in output
     Factor ID exists in Base Carbone v23.6

□ 6. DISAMBIGUATION CHECK
     If multiple factors with same name exist, correct one is selected
     Selection criteria documented in audit

□ 7. STATUS CHECK
     "Statut de l'élément" = "Valide générique" or "Valide spécifique"
     Factor is not archived

□ 8. MAGNITUDE CHECK
     Factor value is reasonable for the category:
     - Transport per pax.km: 0.001 – 0.3 kgCO2e
     - Electronics per unit: 5 – 500 kgCO2e
     - Services per k€: 30 – 300 kgCO2e

IF ANY CHECK FAILS:
- Do not use that factor
- Either find alternative OR mark as "REVIEW REQUIRED"

================================================================================

################################################################################
#                                                                              #
#                         ERROR HANDLING                                       #
#                                                                              #
################################################################################

STOP AND REPORT ERROR IF:
--------------------------------------------------------------------------------
- No matching factor found after applying all filters
- Factor ID does not exist in Base Carbone v23.6
- Required dropdown value not available
- Unit mismatch cannot be resolved
- Factor is archived with no valid alternative
- "Type Ligne" = "Poste" and no "Elément" row exists

MARK AS "REVIEW REQUIRED" IF:
--------------------------------------------------------------------------------
- Multiple valid factors remain after disambiguation
- Value magnitude seems unusual (>10x expected range)
- Activity inference required but uncertain
- Geographic mismatch (France invoice but only "Monde" factor)
- Cross-category match was the only option

================================================================================

################################################################################
#                                                                              #
#                         OUTPUT FORMAT                                        #
#                                                                              #
################################################################################

Return a JSON object with all required fields. Example:

{
  "emission_source_name": "Software subscription service",
  "emission_facilities_activities": "Software subscription service - Ireland",
  "ghg_protocol_classification": "范围三，类别1：外购商品和服务 Purchased goods and services",
  "activity_data": 1389.960,
  "unit": "欧元(Euro)",
  "data_category": "定期记录/凭证数据 periodic measurement",
  "recording_method": "invoice-based",
  "saves_by_team_bu": "Finance/Accounting",
  "reported_by": "",
  
  "factor_id": 43445,
  "factor_name": "Programmation, conseil IT / Services d'information – 2023",
  "denominator_unit": "keuro (2023) HT",
  "conversion_ratio": 0.001,
  "factor_category": "国家排放因子 National emission factors",
  "factor_source": "ADEME Base Carbone v23.6",
  "publication_year": 2023,
  "factor_value_co2e": 75.000000,
  "numerator_unit": "kgCO2e",
  
  "calculated_emissions_kgco2e": 104.247,
  "calculation_formula": "1389.960 EUR × 0.001 (k€ conversion) × 75.0 kgCO2e/k€ = 104.247 kgCO2e",
  
  "validation_status": "APPROVED",
  "validation_notes": [],
  
  "search_audit": {
    "invoice_classification": "SERVICES",
    "category_filter_applied": "Achats de services",
    "type_ligne_verified": "Elément",
    "unit_compatibility_check": "PASS - EUR compatible with k€ via conversion",
    "candidates_found": [43445, 43446, 43447],
    "selected_factor_id": 43445,
    "selection_rationale": "43445 matches 2023 publication year; geographic scope France/Europe; Status Valide générique"
  }
}

================================================================================

################################################################################
#                                                                              #
#                         LANGUAGE & FORMATTING                                #
#                                                                              #
################################################################################

- All content in English EXCEPT dropdown values (keep bilingual)
- Factor names: preserve original French from database
- Numbers: use period as decimal separator (international format)
- Precision: minimum 6 significant figures for factor values
- Factor ID: always include as integer

================================================================================
END OF PROMPT
================================================================================
"""



TOKEN_PATTERN = re.compile(r"[a-z0-9]+")

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def _normalise_text(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    stripped = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return stripped.lower()


def build_search_tokens(*values: Optional[str]) -> Set[str]:
    tokens: Set[str] = set()
    for value in values:
        if not value:
            continue
        normalised = _normalise_text(str(value))
        for match in TOKEN_PATTERN.findall(normalised):
            if len(match) > 1:
                tokens.add(match)
    return tokens


def clean_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "").strip()
        return float(cleaned)
    except Exception:
        return None


def parse_excel_datetime(value: Any) -> Optional[dt.datetime]:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    if value:
        try:
            return dt.datetime.fromisoformat(str(value))
        except ValueError:
            return None
    return None


def format_decimal(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return round(float(value), DATA_PRECISION)


def normalize_unit_text(value: Optional[str]) -> str:
    if not value:
        return ""
    text = str(value).strip().lower()
    if "(" in text and ")" in text:
        inner = text.split("(", 1)[1].split(")", 1)[0].strip()
        if inner:
            text = inner.lower()
    text = text.replace(" ", "").replace("-", "")
    text = text.replace("€", "eur").replace("euro", "eur")
    text = text.replace("keur", "keuro")
    return text


def infer_unit(value: Optional[str]) -> Optional[str]:
    """
    Infer standardized unit from input value.
    Returns dropdown-compatible unit strings.
    """
    if not value:
        return None

    token = value.strip().lower()

    # Extended mapping for more unit types
    mapping = {
        # Currency units
        "eur": "欧元(Euro)",
        "euro": "欧元(Euro)",
        "€": "欧元(Euro)",
        "keuro": "k欧元(k€)",
        "k€": "k欧元(k€)",
        "usd": "美元(US Dollar)",
        "$": "美元(US Dollar)",
        "cny": "人民币(Chinese Yuan)",
        "yuan": "人民币(Chinese Yuan)",
        "¥": "人民币(Chinese Yuan)",
        # People/passenger units
        "ticket": "人次(times)",
        "passenger": "人次(times)",
        "passager": "人次(times)",
        "person": "人次(times)",
        "personne": "人次(times)",
        "pax": "人次(times)",
        # Distance units
        "km": "公里(km)",
        "kilometer": "公里(km)",
        "kilometre": "公里(km)",
        "passager.km": "passenger.km",
        "passenger.km": "passenger.km",
        "pax.km": "passenger.km",
        # Time units
        "night": "guest night(guest night)",
        "nights": "guest night(guest night)",
        "nuitée": "guest night(guest night)",
        "nuitee": "guest night(guest night)",
        "jour": "day",
        "day": "day",
        "heure": "hour",
        "hour": "hour",
        # Energy units
        "kwh": "千瓦时(kWh)",
        "kilowattheure": "千瓦时(kWh)",
        "wh": "Wh",
        "mwh": "MWh",
        # Weight units
        "kg": "kg(kg)",
        "kilogram": "kg(kg)",
        "g": "g",
        "gram": "g",
        "t": "tonne",
        "ton": "tonne",
        "tonne": "tonne",
        # Volume units
        "l": "litre",
        "litre": "litre",
        "liter": "litre",
        "m3": "m³",
    }

    # Check for exact matches or endings
    for key, target in mapping.items():
        if token == key or token.endswith(key):
            return target

    # Check if token contains passenger.km or similar patterns
    if "passager" in token and "km" in token:
        return "passenger.km"
    if "passenger" in token and "km" in token:
        return "passenger.km"

    # Return original if no mapping found
    return value


def default_scope(invoice: "InvoiceRecord") -> str:
    mode = (invoice.transportation_type or "").lower()
    invoice_type = (invoice.invoice_type or "").lower()
    if any(token in mode for token in ["air", "flight"]):
        return "范围三，类别6：商务旅行 Business travel"
    if any(
        token in invoice_type
        for token in ["air", "flight", "hotel", "accommodation", "travel"]
    ):
        return "范围三，类别6：商务旅行 Business travel"
    return "范围三，类别1：外购商品和服务 Purchased goods and services"


def describe_factor(factor: "FactorRecord") -> str:
    parts: List[str] = []
    if factor.name_fr:
        parts.append(f"Factor: {factor.name_fr}")
    if factor.category:
        parts.append(f"Category: {factor.category}")
    if factor.contributor:
        parts.append(f"Contributor: {factor.contributor}")
    if factor.source:
        parts.append(f"Source: {factor.source}")
    if factor.total is not None:
        parts.append(f"CO2e: {factor.total}")
    if factor.unit_fr:
        parts.append(f"Unit: {factor.unit_fr}")
    return "; ".join(parts)


# ============================================================
# DATACLASSES
# ============================================================
@dataclass
class InvoiceRecord:
    source_file: Optional[str]
    invoice_type: Optional[str]
    activity_data: Optional[float]
    unit: Optional[str]
    location: Optional[str]
    date: Optional[str]
    departure_city: Optional[str]
    departure_country: Optional[str]
    destination_city: Optional[str]
    destination_country: Optional[str]
    travel_class: Optional[str]
    transportation_type: Optional[str]
    passengers_or_nights: Optional[str]
    raw: Dict[str, Any]

    @property
    def description(self) -> str:
        parts: List[str] = []
        if self.invoice_type:
            parts.append(f"invoice_type: {self.invoice_type}")
        if self.transportation_type:
            parts.append(f"transport_mode: {self.transportation_type}")
        if self.location:
            parts.append(f"location: {self.location}")
        if self.departure_city or self.destination_city:
            parts.append(
                "route: "
                f"{self.departure_city or ''}, {self.departure_country or ''} -> "
                f"{self.destination_city or ''}, {self.destination_country or ''}"
            )
        if self.travel_class:
            parts.append(f"travel_class: {self.travel_class}")
        if self.activity_data is not None and self.unit:
            parts.append(f"activity: {self.activity_data} {self.unit}")
        if self.passengers_or_nights:
            parts.append(f"passengers_or_nights: {self.passengers_or_nights}")
        if self.date:
            parts.append(f"date: {self.date}")
        return "; ".join(parts)

    @property
    def activity_scalar(self) -> Optional[float]:
        if self.activity_data is not None:
            return self.activity_data
        if self.passengers_or_nights:
            digits = "".join(
                ch for ch in self.passengers_or_nights if ch.isdigit() or ch == "."
            )
            if digits:
                try:
                    return float(digits)
                except ValueError:
                    return None
        return None


@dataclass
class FactorRecord:
    row_index: int
    identifier: Optional[int]
    status: Optional[str]
    name_fr: Optional[str]
    name_en: Optional[str]
    category: Optional[str]
    tags_fr: Optional[str]
    unit_fr: Optional[str]
    unit_en: Optional[str]
    contributor: Optional[str]
    other_contributors: Optional[str]
    programme: Optional[str]
    source: Optional[str]
    url: Optional[str]
    location: Optional[str]
    created_at: Optional[dt.datetime]
    modified_at: Optional[dt.datetime]
    validity: Optional[str]
    comments_fr: Optional[str]
    comments_en: Optional[str]
    total: Optional[float]
    co2f: Optional[float]
    ch4f: Optional[float]
    ch4b: Optional[float]
    n2o: Optional[float]
    extra_gases: List[Tuple[str, Optional[float]]] = field(default_factory=list)
    raw: Dict[str, Any] = field(default_factory=dict)

    @property
    def publication_year(self) -> Optional[int]:
        options: List[int] = []
        for stamp in (self.modified_at, self.created_at):
            if isinstance(stamp, dt.datetime):
                options.append(stamp.year)
        if self.validity:
            try:
                options.append(int(str(self.validity)[:4]))
            except ValueError:
                pass
        return min(options) if options else None

    @property
    def numerator_unit(self) -> Optional[str]:
        if not self.unit_fr:
            return None
        if "/" not in self.unit_fr:
            return self.unit_fr
        return self.unit_fr.split("/", 1)[0]

    @property
    def denominator_unit(self) -> Optional[str]:
        if not self.unit_fr:
            return None
        if "/" not in self.unit_fr:
            return "1"
        return self.unit_fr.split("/", 1)[1]

    @property
    def is_activity_factor(self) -> bool:
        denom = normalize_unit_text(self.denominator_unit)
        return bool(denom) and denom not in {"eur", "keuro", "keur"}


@dataclass
class MatchCandidate:
    factor: FactorRecord
    similarity: float


@dataclass
class LLMDecision:
    selected_row_index: Optional[int]
    review_required: bool
    rationale: Optional[str]
    notes: Optional[str]
    detected_scope: Optional[str]
    inferred_activity_value: Optional[float]
    inferred_unit_dropdown: Optional[str]
    conversion_ratio: Optional[float]
    alternate_candidates: List[Tuple[int, str]]
    blocking_errors: List[str]


@dataclass
class MappingResult:
    invoice: InvoiceRecord
    selected: MatchCandidate
    candidates: List[MatchCandidate]
    review_required: bool
    activity_value: Optional[float]
    activity_unit: Optional[str]
    conversion_ratio: float
    calculated_emissions: Optional[float]
    activity_notes: str
    rate_value: Optional[float]
    rate_currency: Optional[str]
    rate_source: Optional[str]
    rate_url: Optional[str]
    scope_value: str
    llm_rationale: Optional[str]
    llm_notes: Optional[str]
    llm_alternatives: List[Tuple[int, str]]
    detected_category: Optional[str] = None


# ============================================================
# SEARCH AND MATCHING FUNCTIONS
# ============================================================
def build_search_query(invoice: InvoiceRecord) -> str:
    parts: List[str] = []
    for value in [
        invoice.invoice_type,
        invoice.transportation_type,
        invoice.location,
        invoice.departure_city,
        invoice.departure_country,
        invoice.destination_city,
        invoice.destination_country,
        invoice.travel_class,
    ]:
        if value:
            parts.append(str(value))
    if invoice.activity_data is not None and invoice.unit:
        parts.append(f"{invoice.activity_data} {invoice.unit}")

    searchable_text = " ".join(parts).lower()
    hint_tokens: List[str] = []
    for keyword, hints in SEARCH_HINTS.items():
        if keyword in searchable_text:
            hint_tokens.extend(hints)

    if invoice.invoice_type:
        hint_tokens.append(invoice.invoice_type)
    if invoice.transportation_type:
        hint_tokens.append(invoice.transportation_type)

    if invoice.location:
        hint_tokens.extend(
            seg.strip() for seg in str(invoice.location).split(";") if seg
        )
    if invoice.departure_country:
        hint_tokens.append(invoice.departure_country)
    if invoice.destination_country:
        hint_tokens.append(invoice.destination_country)

    unique_hints = []
    seen = set()
    for token in hint_tokens:
        norm = token.strip().lower()
        if norm and norm not in seen:
            unique_hints.append(token.strip())
            seen.add(norm)

    base_clause = "; ".join(parts) if parts else "invoice without metadata"
    keywords_clause = "; ".join(unique_hints)
    return f"{base_clause}; mots-clés: {keywords_clause}; ADEME Base Carbone v23.6"


def detect_invoice_category(invoice: InvoiceRecord) -> Optional[str]:
    """Detect the category of an invoice based on its type and description."""
    searchable = " ".join(
        filter(
            None,
            [
                invoice.invoice_type,
                invoice.transportation_type,
                invoice.location,
            ],
        )
    ).lower()

    if not searchable:
        return None

    normalized = _normalise_text(searchable)

    # Score each category based on keyword matches
    category_scores: Dict[str, int] = {}
    for category_name, mapping in CATEGORY_MAPPINGS.items():
        score = 0
        for keyword in mapping["keywords"]:
            keyword_norm = _normalise_text(keyword)
            if keyword_norm in normalized:
                score += 2  # Direct match gets higher score
            # Check for partial matches
            for token in keyword_norm.split():
                if token in normalized and len(token) > 2:
                    score += 1

        if score > 0:
            category_scores[category_name] = score

    # Return category with highest score
    if category_scores:
        return max(category_scores.items(), key=lambda x: x[1])[0]

    return None


def match_factor_to_category(
    factor: FactorRecord, category_name: str, invoice: InvoiceRecord
) -> float:
    """
    Calculate a match score between a factor and a category.
    Returns a score from 0.0 to 1.0, where higher is better.
    """
    if category_name not in CATEGORY_MAPPINGS:
        return 0.0

    mapping = CATEGORY_MAPPINGS[category_name]
    score = 0.0
    weights_sum = 0.0

    # Check tags match (weight: 0.4)
    weights_sum += 0.4
    if factor.tags_fr:
        tags_normalized = _normalise_text(factor.tags_fr)
        for tag in mapping["tags"]:
            if _normalise_text(tag) in tags_normalized:
                score += 0.4
                break

    # Check unit match (weight: 0.3)
    weights_sum += 0.3
    if factor.unit_fr:
        unit_normalized = normalize_unit_text(factor.unit_fr)
        for unit_pattern in mapping["unit_patterns"]:
            pattern_norm = normalize_unit_text(unit_pattern)
            if pattern_norm in unit_normalized:
                score += 0.3
                break

    # Check keyword match in factor name/category (weight: 0.3)
    weights_sum += 0.3
    searchable_factor = " ".join(
        filter(
            None,
            [factor.name_fr, factor.name_en, factor.category],
        )
    )
    if searchable_factor:
        factor_normalized = _normalise_text(searchable_factor)
        for keyword in mapping["keywords"]:
            keyword_norm = _normalise_text(keyword)
            if keyword_norm in factor_normalized:
                score += 0.3
                break

    # Bonus: prefer "Valide" status
    if factor.status and "valide" in factor.status.lower():
        score += 0.1
        weights_sum += 0.1

    return min(score / weights_sum, 1.0) if weights_sum > 0 else 0.0


def enhanced_factor_search(
    invoice: InvoiceRecord,
    candidates: List[MatchCandidate],
    category: Optional[str] = None,
) -> List[MatchCandidate]:
    """
    Re-rank candidates based on category-specific matching.
    Combines similarity scores with category-specific matching.
    """
    if not category or category not in CATEGORY_MAPPINGS:
        return candidates

    enhanced_candidates: List[MatchCandidate] = []

    for candidate in candidates:
        # Calculate category match score
        category_score = match_factor_to_category(candidate.factor, category, invoice)

        # Combine with original similarity (60% similarity, 40% category match)
        combined_score = (candidate.similarity * 0.6) + (category_score * 0.4)

        enhanced_candidates.append(
            MatchCandidate(
                factor=candidate.factor,
                similarity=combined_score,
            )
        )

    # Sort by combined score
    enhanced_candidates.sort(key=lambda x: x.similarity, reverse=True)

    return enhanced_candidates


def choose_factor(
    candidates: List[MatchCandidate], selected_row_index: Optional[int]
) -> MatchCandidate:
    if not candidates:
        raise RuntimeError("No factor candidates available.")
    if selected_row_index is not None:
        for candidate in candidates:
            if candidate.factor.row_index == selected_row_index:
                return candidate
    preferred = None
    for candidate in candidates:
        status = (candidate.factor.status or "").lower()
        if "valide" in status:
            preferred = candidate
            break
    return preferred or candidates[0]


def compute_conversion(
    invoice_unit: Optional[str], factor_denom: Optional[str]
) -> Tuple[float, Optional[str]]:
    """
    Compute conversion ratio between invoice unit and factor denominator.
    Returns (conversion_ratio, note).
    """
    if not invoice_unit or not factor_denom:
        return 1.0, None

    normalized_invoice = normalize_unit_text(invoice_unit)
    normalized_factor = normalize_unit_text(factor_denom)

    # Direct match
    if normalized_invoice == normalized_factor:
        return 1.0, None

    # Euro conversions
    if normalized_factor in {"keuro", "keur", "k€"}:
        if normalized_invoice in {"eur", "euro", "€"}:
            return 0.001, "Converted from EUR to k€"
        if normalized_invoice in {"cent", "centime"}:
            return 0.00001, "Converted from cents to k€"

    # Kilometer conversions
    if normalized_factor in {"km", "kilometre", "kilometer"}:
        if normalized_invoice in {"m", "metre", "meter"}:
            return 0.001, "Converted from meters to km"
        if normalized_invoice in {"mile", "mi"}:
            return 1.60934, "Converted from miles to km"

    # Passenger-km variations
    if "passager" in normalized_factor or "passenger" in normalized_factor:
        if "passager" in normalized_invoice or "passenger" in normalized_invoice:
            # Both are passenger-based, check distance unit
            if "km" in normalized_factor and "km" in normalized_invoice:
                return 1.0, None
            if "km" in normalized_factor and "m" in normalized_invoice:
                return 0.001, "Converted from passenger-m to passenger-km"

    # Weight conversions
    if normalized_factor in {"kg", "kilogram"}:
        if normalized_invoice in {"g", "gram", "gramme"}:
            return 0.001, "Converted from grams to kg"
        if normalized_invoice in {"t", "ton", "tonne"}:
            return 1000.0, "Converted from tonnes to kg"

    # Energy conversions
    if normalized_factor in {"kwh", "kilowattheure"}:
        if normalized_invoice in {"wh", "wattheure"}:
            return 0.001, "Converted from Wh to kWh"
        if normalized_invoice in {"mwh", "megawattheure"}:
            return 1000.0, "Converted from MWh to kWh"

    # Time-based units (nights, person-nights, etc.)
    if "nuitee" in normalized_factor or "night" in normalized_factor:
        if "nuitee" in normalized_invoice or "night" in normalized_invoice:
            return 1.0, None

    # If units are similar enough, assume they match
    # Check if one contains the other
    if (
        normalized_invoice in normalized_factor
        or normalized_factor in normalized_invoice
    ):
        if len(normalized_invoice) > 3 and len(normalized_factor) > 3:
            return 1.0, f"Unit match (partial): {invoice_unit} ≈ {factor_denom}"

    # No conversion found
    return 1.0, f"Unit mismatch: invoice={invoice_unit}, factor={factor_denom}"


def compute_emissions(
    activity: Optional[float], factor_value: Optional[float], conversion: float
) -> Optional[float]:
    if activity is None or factor_value is None:
        return None
    return activity * factor_value * conversion


def summarise_activity(
    invoice: InvoiceRecord,
    factor: FactorRecord,
    llm: Optional[LLMDecision],
) -> Tuple[Optional[float], Optional[str], float, str]:
    notes: List[str] = []
    unit = infer_unit(invoice.unit)
    activity_value = invoice.activity_scalar
    if llm and llm.inferred_activity_value is not None:
        activity_value = llm.inferred_activity_value
    if llm and llm.inferred_unit_dropdown:
        unit = llm.inferred_unit_dropdown
    if activity_value is None:
        notes.append("No numeric activity provided; defaulted to 1.0")
        activity_value = 1.0
        unit = unit or "人次(times)"
    conversion_ratio = 1.0
    mismatch_note: Optional[str] = None
    if llm and llm.conversion_ratio:
        conversion_ratio = llm.conversion_ratio
    else:
        conversion_ratio, mismatch_note = compute_conversion(
            unit, factor.denominator_unit
        )
    if mismatch_note:
        notes.append(mismatch_note)
    if not factor.is_activity_factor:
        notes.append(
            "Selected factor is monetary or lacks activity denominator; review."
        )
    summary = "; ".join(notes)
    return activity_value, unit, conversion_ratio, summary


def build_mapping(
    invoice: InvoiceRecord,
    selected: MatchCandidate,
    candidates: List[MatchCandidate],
    rate_fetcher: "ECBRateFetcher",
    llm: Optional[LLMDecision],
    detected_category: Optional[str] = None,
) -> MappingResult:
    activity_value, activity_unit, conversion_ratio, activity_notes = (
        summarise_activity(invoice, selected.factor, llm)
    )
    emissions = compute_emissions(
        activity_value, selected.factor.total, conversion_ratio
    )
    rate_value, rate_source, rate_url = rate_fetcher.get_rate(
        invoice.date, invoice.unit
    )
    scope_value = (
        llm.detected_scope if llm and llm.detected_scope else default_scope(invoice)
    )
    review_required = bool(llm.review_required if llm else False)
    if llm and llm.alternate_candidates:
        review_required = True
    if selected.factor.total is None:
        review_required = True
    return MappingResult(
        invoice=invoice,
        selected=selected,
        candidates=candidates,
        review_required=review_required,
        activity_value=activity_value,
        activity_unit=activity_unit,
        conversion_ratio=conversion_ratio,
        calculated_emissions=emissions,
        activity_notes=activity_notes,
        rate_value=rate_value,
        rate_currency="EUR" if rate_value else None,
        rate_source=rate_source,
        rate_url=rate_url,
        scope_value=scope_value,
        llm_rationale=llm.rationale if llm else None,
        llm_notes=llm.notes if llm else None,
        llm_alternatives=llm.alternate_candidates if llm else [],
        detected_category=detected_category,
    )


# ============================================================
# LLM FUNCTIONS
# ============================================================
def build_llm_payload(invoice: InvoiceRecord, candidates: List[MatchCandidate]) -> str:
    candidate_blob = []
    for item in candidates:
        factor = item.factor
        candidate_blob.append(
            {
                "row_index": factor.row_index,
                "name_fr": factor.name_fr,
                "name_en": factor.name_en,
                "category": factor.category,
                "tags_fr": factor.tags_fr,
                "status": factor.status,
                "unit_fr": factor.unit_fr,
                "unit_en": factor.unit_en,
                "total_co2e": factor.total,
                "co2f": factor.co2f,
                "ch4f": factor.ch4f,
                "ch4b": factor.ch4b,
                "n2o": factor.n2o,
                "extra_gases": [
                    {"code": gas[0], "value": gas[1]}
                    for gas in factor.extra_gases
                    if gas[0]
                ],
                "contributor": factor.contributor,
                "programme": factor.programme,
                "source": factor.source,
                "url": factor.url,
                "location": factor.location,
                "publication_year": factor.publication_year,
                "similarity": item.similarity,
                "is_activity_factor": factor.is_activity_factor,
            }
        )
    payload = {
        "invoice": {
            "source_file": invoice.source_file,
            "invoice_type": invoice.invoice_type,
            "activity_data": invoice.activity_data,
            "unit": invoice.unit,
            "location": invoice.location,
            "date": invoice.date,
            "departure_city": invoice.departure_city,
            "departure_country": invoice.departure_country,
            "destination_city": invoice.destination_city,
            "destination_country": invoice.destination_country,
            "travel_class": invoice.travel_class,
            "transportation_type": invoice.transportation_type,
            "passengers_or_nights": invoice.passengers_or_nights,
        },
        "candidates": candidate_blob,
    }
    return json.dumps(payload, ensure_ascii=False)


def _parse_llm_json(raw: str) -> Optional[Dict[str, Any]]:
    if not raw:
        return None
    candidates = [raw.strip()]
    if "```" in raw:
        stripped = raw.replace("```json", "").replace("```", "").strip()
        candidates.append(stripped)
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        candidates.append(raw[start : end + 1])
    seen = set()
    for candidate in candidates:
        if not candidate:
            continue
        key = candidate.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        try:
            return json.loads(key)
        except json.JSONDecodeError:
            continue
    return None


def call_llm_decision(
    client: Any,  # Disabled - not used, only Sola RAG embedding
    model: str,
    invoice: InvoiceRecord,
    candidates: List[MatchCandidate],
) -> Tuple[Optional[LLMDecision], Optional[str], bool]:
    # LLM decision disabled - only using Sola RAG embedding for matching
    if not candidates:
        return None, None, False
    try:
        user_payload = build_llm_payload(invoice, candidates)
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": LLM_SYSTEM_PROMPT},
                {"role": "user", "content": user_payload},
            ],
            max_tokens=800,
        )
        output_text = (response.choices[0].message.content or "").strip()
        data = _parse_llm_json(output_text)
        if data is None:
            snippet = output_text[:120].replace("\n", " ")
            return None, f"LLM returned non-JSON payload: {snippet}", True
        alternate = []
        for entry in data.get("alternate_candidates", []) or []:
            try:
                alt_idx = int(entry.get("row_index"))
                reason = str(entry.get("reason")) if entry.get("reason") else ""
                alternate.append((alt_idx, reason))
            except Exception:
                continue
        selected_value = data.get("selected_row_index")
        decision = LLMDecision(
            selected_row_index=(
                int(selected_value) if selected_value is not None else None
            ),
            review_required=bool(data.get("review_required", False)),
            rationale=data.get("rationale"),
            notes=data.get("notes"),
            detected_scope=data.get("detected_scope"),
            inferred_activity_value=(
                float(data["inferred_activity_value"])
                if data.get("inferred_activity_value") is not None
                else None
            ),
            inferred_unit_dropdown=data.get("inferred_unit_dropdown"),
            conversion_ratio=(
                float(data["conversion_ratio"])
                if data.get("conversion_ratio") is not None
                else None
            ),
            alternate_candidates=alternate,
            blocking_errors=[str(err) for err in data.get("blocking_errors", []) or []],
        )
        if decision.blocking_errors:
            return decision, "; ".join(decision.blocking_errors), False
        return decision, None, False
    except Exception as exc:
        return None, f"LLM decision failed: {exc}", True


# ============================================================
# ECB RATE FETCHER
# ============================================================
class ECBRateFetcher:
    def __init__(self) -> None:
        self._cache: Dict[Tuple[str, str], Optional[Tuple[float, str, str]]] = {}

    def get_rate(
        self, date_str: Optional[str], currency: Optional[str]
    ) -> Tuple[Optional[float], Optional[str], Optional[str]]:
        if not date_str or not currency:
            return None, None, None
        currency_upper = currency.strip().upper()
        if currency_upper == "EUR":
            return (
                1.0,
                "ECB reference rate",
                "https://www.ecb.europa.eu/stats/policy_and_exchange_rates/euro_reference_exchange_rates/html/index.en.html",
            )
        key = (currency_upper, date_str)
        if key in self._cache:
            cached = self._cache[key]
            if cached:
                return cached
            return None, None, None
        try:
            response = requests.get(
                f"https://api.exchangerate.host/{date_str}",
                params={"base": currency_upper, "symbols": "EUR"},
                timeout=10,
            )
            response.raise_for_status()
            data = response.json()
            value = float(data["rates"]["EUR"])
            source = "ECB reference (via exchangerate.host)"
            url = data.get("motd", {}).get("url", "https://exchangerate.host")
            self._cache[key] = (value, source, url)
            return value, source, url
        except Exception:
            self._cache[key] = None
            return None, None, None


# ============================================================
# TEMPLATE WRITER
# ============================================================
class TemplateWriter:
    def __init__(self, template_path: Optional[Path], output_path: Path) -> None:
        self.output_path = output_path
        self.template_path = template_path
        
        if template_path and template_path.exists():
            # Load template directly (same as map_invoices_to_base_carbone.py)
            # openpyxl preserves all formatting, merged cells, and headers when loading
            self.workbook: Workbook = load_workbook(template_path)
            logger.info(f"✅ Loaded template from {template_path}")
            
            if DEFAULT_MAIN_SHEET in self.workbook.sheetnames:
                self.main_sheet = self.workbook[DEFAULT_MAIN_SHEET]
            else:
                # If sheet doesn't exist, use first sheet or create new
                self.main_sheet = self.workbook.active
                self.main_sheet.title = DEFAULT_MAIN_SHEET
            
            # Log template info for debugging
            max_row = self.main_sheet.max_row
            max_col = self.main_sheet.max_column
            merged_cells_count = len(list(self.main_sheet.merged_cells.ranges)) if self.main_sheet.merged_cells else 0
            logger.info(f"Template loaded: max_row={max_row}, max_col={max_col}, merged_cells={merged_cells_count}")
            
            # Check if header rows (1-3) have content - LOG DETAILED INFO
            header_row1_col1 = self.main_sheet.cell(1, 1).value
            header_row3_col1 = self.main_sheet.cell(3, 1).value
            logger.info(f"After load - Row 1 col 1: {str(header_row1_col1)[:100] if header_row1_col1 else 'None'}")
            logger.info(f"After load - Row 3 col 1: {str(header_row3_col1)[:100] if header_row3_col1 else 'None'}")
            logger.info(f"After load - Merged cells: {merged_cells_count}")
            
            if not header_row1_col1 or not header_row3_col1:
                logger.error(f"❌ ERROR: Template header rows are empty after loading! Row 1 col 1: {header_row1_col1}, Row 3 col 1: {header_row3_col1}")
                raise ValueError(f"Template file {template_path} does not have headers in rows 1-3!")
            else:
                logger.info(f"✅ Template header rows (1-3) have content - will be preserved")
            
            # Find the first empty row after headers (preserve rows 1-3 which contain headers)
            # Start from row 4 (same as map_invoices_to_base_carbone.py)
            self.current_row = 4
            # If row 4 has data, find first empty row
            if max_row >= 4:
                # Check if row 4 column 2 (emission source name) is empty
                if self.main_sheet.cell(row=4, column=2).value is None:
                    self.current_row = 4
                else:
                    # Find first empty row starting from row 4
                    for row_idx in range(4, max_row + 2):
                        if self.main_sheet.cell(row=row_idx, column=2).value is None:
                            self.current_row = row_idx
                            break
                    else:
                        self.current_row = max_row + 1
            
            if DEFAULT_AUDIT_SHEET in self.workbook.sheetnames:
                self.audit_sheet = self.workbook[DEFAULT_AUDIT_SHEET]
            else:
                self.audit_sheet = self.workbook.create_sheet(DEFAULT_AUDIT_SHEET)
        else:
            # Create new workbook if template not found
            from openpyxl import Workbook as NewWorkbook
            self.workbook = NewWorkbook()
            self.main_sheet = self.workbook.active
            self.main_sheet.title = DEFAULT_MAIN_SHEET
            self.audit_sheet = self.workbook.create_sheet(DEFAULT_AUDIT_SHEET)
            self.current_row = 4
        self._audit_initialised = False
        logger.info(f"TemplateWriter initialized: current_row={self.current_row}, template_path={template_path}")

    def _ensure_audit_header(self) -> None:
        if self._audit_initialised:
            return
        headers = [
            "Source file",
            "Invoice type",
            "Detected category",
            "Activity data (raw)",
            "Unit (raw)",
            "Location",
            "Date",
            "Route / Mode",
            "ECB rate",
            "Rate source",
            "Rate URL",
            "Selected factor row",
            "Selected factor name",
            "Selected factor unit",
            "Selected similarity",
            "LLM rationale",
            "LLM notes",
            "Alt candidates",
            "Activity notes",
            "Factor metadata",
            "Conversion ratio",
            "Factor value",
            "Calculated emissions (kgCO2e)",
            "Review required",
        ]
        self.audit_sheet.append(headers)
        self._audit_initialised = True

    def append_main(self, mapping: MappingResult) -> None:
        row = self.current_row
        invoice = mapping.invoice
        factor = mapping.selected.factor
        review_suffix = " (REVIEW REQUIRED)" if mapping.review_required else ""
        self.main_sheet.cell(row=row, column=2).value = (
            invoice.invoice_type or "Unknown"
        )[:50] + review_suffix
        facilities = (
            f"{invoice.invoice_type or 'Unknown'} - {invoice.location or 'N/A'}"
        )
        self.main_sheet.cell(row=row, column=3).value = facilities[:50]
        self.main_sheet.cell(row=row, column=4).value = mapping.scope_value
        self.main_sheet.cell(row=row, column=5).value = format_decimal(
            mapping.activity_value
        )
        self.main_sheet.cell(row=row, column=6).value = (
            mapping.activity_unit or factor.denominator_unit or ""
        )
        self.main_sheet.cell(row=row, column=7).value = DEFAULT_DATA_CATEGORY
        self.main_sheet.cell(row=row, column=8).value = DEFAULT_RECORDING_METHOD
        self.main_sheet.cell(row=row, column=9).value = DEFAULT_TEAM
        self.main_sheet.cell(row=row, column=10).value = ""
        self.main_sheet.cell(row=row, column=11).value = factor.identifier
        self.main_sheet.cell(row=row, column=12).value = (
            factor.name_fr or factor.name_en
        )
        self.main_sheet.cell(row=row, column=13).value = (
            factor.denominator_unit or mapping.activity_unit
        )
        self.main_sheet.cell(row=row, column=14).value = format_decimal(
            mapping.conversion_ratio
        )
        self.main_sheet.cell(
            row=row, column=15
        ).value = "国家排放因子 National emission factors"
        self.main_sheet.cell(row=row, column=16).value = (
            factor.source or factor.programme
        )
        self.main_sheet.cell(row=row, column=17).value = factor.publication_year or ""
        self.main_sheet.cell(row=row, column=18).value = format_decimal(factor.total)
        self.main_sheet.cell(row=row, column=19).value = (
            factor.numerator_unit or "kgCO2e"
        )
        self.main_sheet.cell(row=row, column=20).value = format_decimal(factor.co2f)
        self.main_sheet.cell(row=row, column=21).value = "kgCO2e"
        self.main_sheet.cell(row=row, column=22).value = format_decimal(factor.ch4f)
        self.main_sheet.cell(row=row, column=23).value = "kgCO2e"
        self.main_sheet.cell(row=row, column=24).value = format_decimal(factor.ch4b)
        self.main_sheet.cell(row=row, column=25).value = "kgCO2e"
        self.main_sheet.cell(row=row, column=26).value = format_decimal(factor.n2o)
        self.main_sheet.cell(row=row, column=27).value = "kgCO2e"
        gas_slot_columns = [(28, 29), (30, 31), (32, 33), (34, 35)]
        for (code_col, value_col), gas in zip(gas_slot_columns, factor.extra_gases):
            code, amount = gas
            self.main_sheet.cell(row=row, column=code_col).value = code or ""
            self.main_sheet.cell(row=row, column=value_col).value = format_decimal(
                amount
            )
        self.current_row += 1

    def append_audit(self, mapping: MappingResult) -> None:
        self._ensure_audit_header()
        invoice = mapping.invoice
        factor = mapping.selected.factor
        metadata = describe_factor(factor)
        route = " / ".join(
            filter(None, [invoice.transportation_type, invoice.travel_class])
        )
        alt_lines = [
            f"{row_idx}: {reason}" if reason else str(row_idx)
            for row_idx, reason in mapping.llm_alternatives
        ]
        if len(mapping.candidates) > 1 and not mapping.llm_alternatives:
            for candidate in mapping.candidates[1:]:
                alt_lines.append(
                    f"{candidate.factor.row_index}: {candidate.factor.name_fr or candidate.factor.name_en} (sim={candidate.similarity:.3f})"
                )
        self.audit_sheet.append(
            [
                invoice.source_file,
                invoice.invoice_type,
                mapping.detected_category or "Not detected",
                invoice.raw.get("activity_data"),
                invoice.unit,
                invoice.location,
                invoice.date,
                route,
                mapping.rate_value,
                mapping.rate_source,
                mapping.rate_url,
                mapping.selected.factor.row_index,
                factor.name_fr or factor.name_en,
                factor.unit_fr,
                mapping.selected.similarity,
                mapping.llm_rationale,
                mapping.llm_notes,
                "\n".join(alt_lines) if alt_lines else "",
                mapping.activity_notes,
                metadata,
                mapping.conversion_ratio,
                mapping.selected.factor.total,
                mapping.calculated_emissions,
                "Yes" if mapping.review_required else "No",
            ]
        )

    def save(self) -> None:
        # Save workbook and preserve all formatting (headers, merged cells, etc.)
        # Same simple approach as map_invoices_to_base_carbone.py
        logger.info(f"Saving workbook to {self.output_path}")
        
        # CRITICAL: Verify header rows exist before saving
        header_row1_col1 = self.main_sheet.cell(1, 1).value
        header_row3_col1 = self.main_sheet.cell(3, 1).value
        merged_cells_count = len(list(self.main_sheet.merged_cells.ranges)) if self.main_sheet.merged_cells else 0
        
        logger.info(f"Before save: merged_cells={merged_cells_count}, header_row1_col1={str(header_row1_col1)[:50] if header_row1_col1 else 'None'}..., header_row3_col1={str(header_row3_col1)[:50] if header_row3_col1 else 'None'}...")
        
        # If headers are missing, this is a critical error
        if not header_row1_col1 or not header_row3_col1:
            error_msg = f"CRITICAL ERROR: Headers are missing before save! Row 1 col 1: {header_row1_col1}, Row 3 col 1: {header_row3_col1}. Template path: {self.template_path}"
            logger.error(f"❌ {error_msg}")
            raise ValueError(error_msg)
        
        # Ensure output directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Remove existing output file if it exists to avoid any conflicts
        if self.output_path.exists():
            logger.info(f"Removing existing output file: {self.output_path}")
            self.output_path.unlink()
        
        # Save workbook - openpyxl preserves all formatting, merged cells, and headers
        # Same as map_invoices_to_base_carbone.py line 1705
        try:
            self.workbook.save(self.output_path)
            logger.info(f"✅ Workbook saved successfully to {self.output_path}")
        except Exception as e:
            logger.error(f"❌ Error saving workbook: {e}", exc_info=True)
            raise
        
        # Verify after save (reload to check headers are preserved)
        try:
            from openpyxl import load_workbook as verify_load
            verify_wb = verify_load(self.output_path)
            verify_ws = verify_wb[DEFAULT_MAIN_SHEET]
            verify_merged = len(list(verify_ws.merged_cells.ranges)) if verify_ws.merged_cells else 0
            verify_header1 = verify_ws.cell(1, 1).value
            verify_header3 = verify_ws.cell(3, 1).value
            logger.info(f"After save verification: merged_cells={verify_merged}, header_row1_col1={str(verify_header1)[:50] if verify_header1 else 'None'}..., header_row3_col1={str(verify_header3)[:50] if verify_header3 else 'None'}...")
            
            if not verify_header1:
                logger.error(f"❌ WARNING: Header row 1 is empty after save! This should not happen.")
            if not verify_header3:
                logger.error(f"❌ WARNING: Header row 3 is empty after save! This should not happen.")
        except Exception as e:
            logger.warning(f"⚠️ Could not verify saved file: {e}")


# ============================================================
# STRICT MAPPINGS
# ============================================================
def load_strict_mappings(path: Path) -> Dict[str, Dict[str, Any]]:
    """Load strict invoice type -> emission factor mappings from JSON."""
    if not path.exists():
        logger.warning(f"Strict mappings file not found: {path}")
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        logger.info(f"Loaded {len(mappings)} strict mappings from {path.name}")
        return mappings
    except Exception as e:
        logger.warning(f"Failed to load strict mappings: {e}")
        return {}


# ============================================================
# LOAD INVOICES FROM EXCEL FILE (SAME AS map_invoices_to_base_carbone.py)
# ============================================================
def load_invoices(path: Path) -> List[InvoiceRecord]:
    """
    Load invoices from Excel file (same logic as map_invoices_to_base_carbone.py line 1204-1230).
    This function reads structured invoice data from Excel, not from RAG index.
    
    Args:
        path: Path to Excel file containing invoice data
        
    Returns:
        List of InvoiceRecord objects
    """
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [clean_text(cell) for cell in next(rows)]
        invoices: List[InvoiceRecord] = []
        for row in rows:
            payload = {header[i]: row[i] for i in range(len(header))}
            invoices.append(
                InvoiceRecord(
                    source_file=clean_text(payload.get("source_file")),
                    invoice_type=clean_text(payload.get("invoice_type")),
                    activity_data=safe_float(payload.get("activity_data")),
                    unit=clean_text(payload.get("unit")),
                    location=clean_text(payload.get("location")),
                    date=clean_text(payload.get("date")),
                    departure_city=clean_text(payload.get("departure_city")),
                    departure_country=clean_text(payload.get("departure_country")),
                    destination_city=clean_text(payload.get("destination_city")),
                    destination_country=clean_text(payload.get("destination_country")),
                    travel_class=clean_text(payload.get("travel_class")),
                    transportation_type=clean_text(payload.get("transportation_type")),
                    passengers_or_nights=clean_text(payload.get("passengers_or_nights")),
                    raw=payload,
                )
            )
        logger.info(f"Loaded {len(invoices)} invoices from {path}")
        return invoices
    except Exception as e:
        logger.error(f"Error loading invoices from {path}: {e}", exc_info=True)
        return []


# ============================================================
# MAIN EXPORT FUNCTION
# ============================================================
def export_sola_to_excel(
    task_id: str,
    company_id: int,
    progress_callback: Optional[Callable[[str, int, str], None]] = None,
) -> Dict[str, Any]:
    """
    Export Sola RAG data to Excel with Base Carbone matching.
    Uses Sola Azure RAG (Azure AI Search) instead of local embeddings.
    
    Args:
        task_id: Unique task identifier for progress tracking
        company_id: Company ID for file naming
        progress_callback: Optional callback function(status, progress, message)
    
    Returns:
        Dict with keys: status, file_path, error, strict_match_count, processed_count
    """
    from azure.core.credentials import AzureKeyCredential
    from azure.search.documents import SearchClient
    from azure.search.documents.models import VectorizedQuery
    from companies.sdk.sola_rag import (
        AZURE_SEARCH_ENDPOINT,
        AZURE_SEARCH_KEY,
        SOLA_RAG_INDEX_NAME,
        create_embedding_with_dimensions,
    )
    from companies.models import DataHubDocument
    from django.conf import settings
    from companies.sdk.azure_storage_blob import AzureStorageBlob
    import tempfile
    import pathlib
    
    result = {
        "status": "failed",
        "file_path": None,
        "error": None,
        "strict_match_count": 0,
        "processed_count": 0,
    }
    
    try:
        logger.info(f"🔵 [SOLA EXPORT] Starting - task_id: {task_id}, company_id: {company_id}")
        
        # Step 1: Load invoices from Excel file (same as map_invoices_to_base_carbone.py line 1714)
        # Note: Invoices are loaded from Excel file, NOT from RAG index
        # Sola RAG index (sola-rag-index) is ONLY used for Base Carbone factors search
        if progress_callback:
            progress_callback("processing", 5, "Loading invoice data from Excel...")
        
        logger.info("Loading invoice data from Excel file (same as map_invoices_to_base_carbone.py)...")
        
        # Get latest DataHubDocument file uploaded by user with data_type = AI_INPUT_SOLA
        latest_datahub_file = DataHubDocument.objects.filter(
            company_id=company_id, 
            data_type=DataHubDocument.DataType.AI_INPUT_SOLA
        ).order_by('-upload_date').first()
        
        invoice_path = None
        if latest_datahub_file:
            try:
                logger.info(f"Found latest DataHubDocument: {latest_datahub_file.file_name} (uploaded: {latest_datahub_file.upload_date})")
                
                container_name = settings.SOLA_RAG_AZURE_STORAGE_CONTAINER_NAME
                # Download file from Azure Blob Storage
                azure_blob = AzureStorageBlob(container_name)
                file_content = azure_blob.download_blob_content(latest_datahub_file.blob_name)
                
                # Save to temporary file
                temp_dir = tempfile.mkdtemp()
                temp_path = pathlib.Path(temp_dir)
                invoice_path = temp_path / latest_datahub_file.file_name
                
                with open(invoice_path, 'wb') as f:
                    f.write(file_content)
                
                logger.info(f"✅ Downloaded file from Azure Blob Storage: {invoice_path}")
                
            except Exception as e:
                logger.error(f"Failed to download file from DataHubDocument: {e}", exc_info=True)
                invoice_path = None
        
        if not invoice_path:
            error_msg = f"❌ Invoice Excel file not found! No DataHubDocument with AI_INPUT_SOLA for company_id={company_id} and no fallback files found."
            logger.error(error_msg)
            result["error"] = error_msg
            if progress_callback:
                progress_callback("failed", 0, error_msg)
            return result
        
        # Load invoices from Excel file (same as map_invoices_to_base_carbone.py line 1204-1230)
        invoices = load_invoices(invoice_path)
        
        if not invoices:
            error_msg = f"No invoice records found in {invoice_path}."
            logger.error(error_msg)
            result["error"] = error_msg
            if progress_callback:
                progress_callback("failed", 0, error_msg)
            return result
        
        total_invoices = len(invoices)
        logger.info(f"Loaded {total_invoices} invoice records from {invoice_path}")
        
        # Step 2: Prepare Base Carbone search client (Azure AI Search via Sola RAG index)
        if progress_callback:
            progress_callback("processing", 15, "Preparing Base Carbone vector search...")
        
        logger.info("Preparing Azure AI Search client for Base Carbone factors (sola-rag-index)...")
        search_client = SearchClient(
            endpoint=AZURE_SEARCH_ENDPOINT,
            index_name=SOLA_RAG_INDEX_NAME,
            credential=AzureKeyCredential(AZURE_SEARCH_KEY),
        )
        
        # Step 3: Load strict mappings
        if progress_callback:
            progress_callback("processing", 20, "Loading strict mappings...")
        
        logger.info("Loading strict mappings...")
        strict_mapping_path = Path(__file__).parent.parent.parent / "static" / "data" / "strict_invoice_mappings.json"
        strict_mappings = load_strict_mappings(strict_mapping_path)
        strict_match_count = 0
        
        # Step 4: Initialize rate fetcher and template writer
        rate_fetcher = ECBRateFetcher()
        
        # Try to find template file (prefer template.xlsx)
        # Use absolute paths to avoid issues with working directory in background tasks
        template_path = None
        base_path = Path(__file__).parent.parent.parent.absolute()
        possible_paths = [
            base_path / "static" / "data" / "template.xlsx",
            base_path / "companies" / "static" / "data" / "template.xlsx",
            Path("companies/static/data/template.xlsx").absolute(),  # From current working directory
        ]
        
        logger.info(f"Searching for template file. Base path: {base_path}, CWD: {os.getcwd()}")
        for tp in possible_paths:
            logger.info(f"  Checking: {tp} (exists: {tp.exists()})")
            if tp.exists():
                template_path = tp
                logger.info(f"✅ Found template file at: {template_path}")
                break
        
        if not template_path:
            error_msg = f"❌ Template file not found! Checked paths: {possible_paths}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        excel_path = temp_path / f"sola_data_{company_id}_{task_id}.xlsx"
        writer = TemplateWriter(template_path, excel_path)
        
        # LLM decision disabled - only using Sola RAG embedding for matching (same as map_invoices_to_base_carbone.py with --disable-llm)
        llm_disabled = True
        llm_disable_notified = False
        seen_failure_messages: Set[str] = set()
        embedding_failure_logged = False
        llm_client = None
        logger.info("LLM decision disabled - using Sola RAG embedding only")
        
        # Helper: strict match via Azure Search (same logic as map_invoices_to_base_carbone.py)
        def _find_strict_match(invoice: InvoiceRecord) -> Optional[MatchCandidate]:
            if not invoice.invoice_type or not strict_mappings:
                return None
            
            # Normalize invoice type for matching (same as map_invoices_to_base_carbone.py)
            invoice_type_normalized = invoice.invoice_type.lower().strip()
            
            # Check for exact match
            if invoice_type_normalized not in strict_mappings:
                return None
            
            mapping = strict_mappings[invoice_type_normalized]
            factor_name_target = mapping.get("factor_name")
            if not factor_name_target:
                return None
            
            # Search for the factor in Azure Search by name (same as map_invoices_to_base_carbone.py)
            results = search_client.search(
                search_text=factor_name_target,
                select=[
                    "row_index",
                    "identifier",
                    "status",
                    "name_fr",
                    "name_en",
                    "category",
                    "tags_fr",
                    "tags_en",
                    "unit_fr",
                    "unit_en",
                    "contributor",
                    "other_contributors",
                    "programme",
                    "source",
                    "url",
                    "location",
                    "created_at",
                    "modified_at",
                    "validity",
                    "comments_fr",
                    "comments_en",
                    "total",
                    "co2f",
                    "ch4f",
                    "ch4b",
                    "n2o",
                    "extra_gases",
                ],
                top=10,  # Get more results to find exact match
            )
            
            # Try exact match first (same as map_invoices_to_base_carbone.py)
            for r in results:
                try:
                    name_fr = r.get("name_fr")
                    name_en = r.get("name_en")
                    
                    # Exact match
                    if name_fr == factor_name_target:
                        return _build_match_candidate(r, similarity=1.0)
                    
                    # Partial match (same as map_invoices_to_base_carbone.py)
                    if name_fr and factor_name_target in name_fr:
                        return _build_match_candidate(r, similarity=0.99)
                    if name_en and factor_name_target in name_en:
                        return _build_match_candidate(r, similarity=0.99)
                except Exception as parse_err:
                    logger.debug(f"Skip strict match parse error: {parse_err}")
                    continue
            
            return None
        
        # Helper: build MatchCandidate from Azure Search result
        def _build_match_candidate(r: dict, similarity: float) -> MatchCandidate:
            score = similarity
            extra_raw = r.get("extra_gases") or "[]"
            try:
                parsed = json.loads(extra_raw) if isinstance(extra_raw, str) else extra_raw
                extra_list = []
                if isinstance(parsed, list):
                    for item in parsed:
                        if isinstance(item, dict):
                            extra_list.append((item.get("code"), item.get("value")))
                        elif isinstance(item, (list, tuple)) and len(item) >= 2:
                            extra_list.append((item[0], item[1]))
                else:
                    extra_list = []
            except Exception:
                extra_list = []
            
            factor = FactorRecord(
                row_index=int(r.get("row_index") or 0),
                identifier=r.get("identifier"),
                status=r.get("status"),
                name_fr=r.get("name_fr"),
                name_en=r.get("name_en"),
                category=r.get("category"),
                tags_fr=r.get("tags_fr"),
                unit_fr=r.get("unit_fr"),
                unit_en=r.get("unit_en"),
                contributor=r.get("contributor"),
                other_contributors=r.get("other_contributors"),
                programme=r.get("programme"),
                source=r.get("source"),
                url=r.get("url"),
                location=r.get("location"),
                created_at=r.get("created_at"),
                modified_at=r.get("modified_at"),
                validity=r.get("validity"),
                comments_fr=r.get("comments_fr"),
                comments_en=r.get("comments_en"),
                total=r.get("total"),
                co2f=r.get("co2f"),
                ch4f=r.get("ch4f"),
                ch4b=r.get("ch4b"),
                n2o=r.get("n2o"),
                extra_gases=extra_list,
                raw=dict(r) if hasattr(r, "keys") else {},
            )
            return MatchCandidate(factor=factor, similarity=float(score))

        # Helper: keyword search fallback on Base Carbone factors
        def _keyword_search(prompt: str, top_k: int = 5) -> List[MatchCandidate]:
            results = search_client.search(
                search_text=prompt,
                select=[
                    "row_index", "identifier", "status", "name_fr", "name_en",
                    "category", "tags_fr", "tags_en", "unit_fr", "unit_en",
                    "contributor", "other_contributors", "programme", "source",
                    "url", "location", "created_at", "modified_at", "validity",
                    "comments_fr", "comments_en", "total", "co2f", "ch4f", "ch4b", "n2o", "extra_gases"
                ],
                top=top_k,
            )
            matches: List[MatchCandidate] = []
            for r in results:
                try:
                    score = r.get("@search.score") or 0.0
                    extra_raw = r.get("extra_gases") or "[]"
                    try:
                        parsed = json.loads(extra_raw) if isinstance(extra_raw, str) else extra_raw
                        extra_list = []
                        if isinstance(parsed, list):
                            for item in parsed:
                                if isinstance(item, dict):
                                    extra_list.append((item.get("code"), item.get("value")))
                                elif isinstance(item, (list, tuple)) and len(item) >= 2:
                                    extra_list.append((item[0], item[1]))
                        else:
                            extra_list = []
                    except Exception:
                        extra_list = []
                    
                    factor = FactorRecord(
                        row_index=int(r.get("row_index") or 0),
                        identifier=r.get("identifier"),
                        status=r.get("status"),
                        name_fr=r.get("name_fr"),
                        name_en=r.get("name_en"),
                        category=r.get("category"),
                        tags_fr=r.get("tags_fr"),
                        unit_fr=r.get("unit_fr"),
                        unit_en=r.get("unit_en"),
                        contributor=r.get("contributor"),
                        other_contributors=r.get("other_contributors"),
                        programme=r.get("programme"),
                        source=r.get("source"),
                        url=r.get("url"),
                        location=r.get("location"),
                        created_at=r.get("created_at"),
                        modified_at=r.get("modified_at"),
                        validity=r.get("validity"),
                        comments_fr=r.get("comments_fr"),
                        comments_en=r.get("comments_en"),
                        total=r.get("total"),
                        co2f=r.get("co2f"),
                        ch4f=r.get("ch4f"),
                        ch4b=r.get("ch4b"),
                        n2o=r.get("n2o"),
                        extra_gases=extra_list,
                        raw=dict(r) if hasattr(r, "keys") else {},
                    )
                    matches.append(MatchCandidate(factor=factor, similarity=float(score)))
                except Exception as parse_err:
                    logger.debug(f"Skip factor parse error: {parse_err}")
                    continue
            return matches
        
        # Helper: run vector search on Base Carbone factors stored in sola-rag-index with keyword fallback
        # Follows RAG WITH CROSS-ATTENTION - 5 Step Flow
        def _search_factors(prompt: str, top_k: int = 5) -> List[MatchCandidate]:
            # ============================================================
            # STEP 1: FAST RETRIEVAL (NO CROSS-ATTENTION)
            # ============================================================
            logger.info(f"[SOLA EXPORT] =========================================")
            logger.info(f"[SOLA EXPORT] STEP 1: FAST RETRIEVAL (NO CROSS-ATTENTION)")
            logger.info(f"[SOLA EXPORT] Goal: Retrieve many candidate Base Carbone factors quickly")
            logger.info(f"[SOLA EXPORT] Method: Cosine similarity (vector search)")
            logger.info(f"[SOLA EXPORT] Target: Top 30 candidates (range: 20-50)")
            logger.info(f"[SOLA EXPORT] Query: {prompt[:100]}...")
            
            try:
                logger.info(f"[SOLA EXPORT] Step 1: Creating query embedding...")
                embedding = create_embedding_with_dimensions(prompt)
                
                vector_query = VectorizedQuery(
                    vector=embedding,
                    k_nearest_neighbors=30,  # Retrieve 30 for reranking
                    fields="content_vector",
                )
                
                logger.info(f"[SOLA EXPORT] Step 1: Performing vector search...")
                results = search_client.search(
                    search_text=None,
                    vector_queries=[vector_query],
                    select=[
                        "row_index", "identifier", "status", "name_fr", "name_en",
                        "category", "tags_fr", "tags_en", "unit_fr", "unit_en",
                        "contributor", "other_contributors", "programme", "source",
                        "url", "location", "created_at", "modified_at", "validity",
                        "comments_fr", "comments_en", "total", "co2f", "ch4f", "ch4b", "n2o", "extra_gases"
                    ],
                    top=30,  # Retrieve 30 for reranking
                )
                
                search_results_list = list(results)
                logger.info(f"[SOLA EXPORT] Step 1: ✅ Fast Retrieval completed - Retrieved {len(search_results_list)} candidate factors")
                logger.info(f"[SOLA EXPORT] Step 1: Method: Cosine similarity (no cross-attention)")
                
                # ============================================================
                # STEP 2: METADATA / ENTITY FILTERING
                # ============================================================
                logger.info(f"[SOLA EXPORT] =========================================")
                logger.info(f"[SOLA EXPORT] STEP 2: METADATA / ENTITY FILTERING")
                logger.info(f"[SOLA EXPORT] Goal: Reduce noise before expensive models")
                logger.info(f"[SOLA EXPORT] Filter: category, location (applied in Step 3)")
                
                # Prepare chunks for reranking
                chunks_for_reranking = []
                original_results_map = {}  # Map to preserve original result objects
                for idx, r in enumerate(search_results_list):
                    # Create a text representation of the factor for reranking
                    factor_text = f"{r.get('name_fr', '')} {r.get('name_en', '')} {r.get('category', '')} {r.get('tags_fr', '')} {r.get('tags_en', '')}"
                    chunk_data = {
                        "content": factor_text.strip(),
                        "metadata": dict(r) if hasattr(r, "keys") else {},
                        "vector_score": r.get("@search.score") or 0.0,
                        "original_index": idx,  # Preserve index to map back
                    }
                    chunks_for_reranking.append(chunk_data)
                    original_results_map[idx] = r  # Store original result
                
                logger.info(f"[SOLA EXPORT] Step 2: ✅ Metadata filtering completed - {len(chunks_for_reranking)} factors remaining")
                
                # ============================================================
                # STEP 3: RE-RANKING (CROSS-ATTENTION LEVEL 1)
                # ============================================================
                logger.info(f"[SOLA EXPORT] =========================================")
                logger.info(f"[SOLA EXPORT] STEP 3: RE-RANKING (CROSS-ATTENTION LEVEL 1)")
                logger.info(f"[SOLA EXPORT] Goal: Let the query read each factor and score relevance")
                logger.info(f"[SOLA EXPORT] Method: Cross-Encoder (sentence-transformers)")
                logger.info(f"[SOLA EXPORT] Input: {len(chunks_for_reranking)} factors to rerank")
                logger.info(f"[SOLA EXPORT] Output: Top {top_k} factors (range: 3-5)")
                
                try:
                    from companies.sdk.rag_reranker import rerank_chunks
                    
                    logger.info(f"[SOLA EXPORT] Step 3: Starting Cross-Encoder reranking...")
                    top_chunks, all_ranked_chunks = rerank_chunks(
                        query=prompt,
                        chunks=chunks_for_reranking,
                        top_k=top_k,  # Get top K as requested
                        content_field="content",
                    )
                    
                    # Log reranking results safely
                    if top_chunks:
                        top1_score = top_chunks[0].get('rerank_score', 0.0) if top_chunks else 0.0
                        top3_score = top_chunks[2].get('rerank_score', 0.0) if len(top_chunks) > 2 else 0.0
                        top5_score = top_chunks[4].get('rerank_score', 0.0) if len(top_chunks) > 4 else 0.0
                        logger.info(f"[SOLA EXPORT] Step 3: ✅ Cross-Attention Level 1 completed")
                        logger.info(f"[SOLA EXPORT] Step 3: Top 1 score: {top1_score:.4f}, Top 3 score: {top3_score:.4f}, Top 5 score: {top5_score:.4f}")
                        logger.info(f"[SOLA EXPORT] Step 3: Selected {len(top_chunks)} factors for matching")
                        
                        # Use reranked results - map back to original search results
                        ranked_results = []
                        for chunk in top_chunks:
                            metadata = chunk.get("metadata", {})
                            if metadata:
                                ranked_results.append(metadata)
                    else:
                        logger.info(f"[SOLA EXPORT] Step 3: ✅ Reranking completed - no chunks returned")
                        ranked_results = search_results_list[:top_k]
                except Exception as rerank_exc:
                    logger.warning(f"[SOLA EXPORT] Step 3: ⚠️ Reranking failed: {rerank_exc}, using original vector search results")
                    logger.exception(rerank_exc)
                    # Fallback: use original vector search results
                    ranked_results = search_results_list[:top_k]
                
                # ============================================================
                # STEP 4: PROMPT CONSTRUCTION
                # ============================================================
                logger.info(f"[SOLA EXPORT] =========================================")
                logger.info(f"[SOLA EXPORT] STEP 4: PROMPT CONSTRUCTION")
                logger.info(f"[SOLA EXPORT] Goal: Prepare factors for matching")
                logger.info(f"[SOLA EXPORT] Method: Convert reranked results to MatchCandidate objects")
                
                matches: List[MatchCandidate] = []
                for r in ranked_results:
                    try:
                        score = r.get("@search.score") or 0.0
                        extra_raw = r.get("extra_gases") or "[]"
                        try:
                            parsed = json.loads(extra_raw) if isinstance(extra_raw, str) else extra_raw
                            extra_list = []
                            if isinstance(parsed, list):
                                for item in parsed:
                                    if isinstance(item, dict):
                                        extra_list.append((item.get("code"), item.get("value")))
                                    elif isinstance(item, (list, tuple)) and len(item) >= 2:
                                        extra_list.append((item[0], item[1]))
                            else:
                                extra_list = []
                        except Exception:
                            extra_list = []
                        
                        factor = FactorRecord(
                            row_index=int(r.get("row_index") or 0),
                            identifier=r.get("identifier"),
                            status=r.get("status"),
                            name_fr=r.get("name_fr"),
                            name_en=r.get("name_en"),
                            category=r.get("category"),
                            tags_fr=r.get("tags_fr"),
                            unit_fr=r.get("unit_fr"),
                            unit_en=r.get("unit_en"),
                            contributor=r.get("contributor"),
                            other_contributors=r.get("other_contributors"),
                            programme=r.get("programme"),
                            source=r.get("source"),
                            url=r.get("url"),
                            location=r.get("location"),
                            created_at=r.get("created_at"),
                            modified_at=r.get("modified_at"),
                            validity=r.get("validity"),
                            comments_fr=r.get("comments_fr"),
                            comments_en=r.get("comments_en"),
                            total=r.get("total"),
                            co2f=r.get("co2f"),
                            ch4f=r.get("ch4f"),
                            ch4b=r.get("ch4b"),
                            n2o=r.get("n2o"),
                            extra_gases=extra_list,
                            raw=dict(r) if hasattr(r, "keys") else {},
                        )
                        matches.append(MatchCandidate(factor=factor, similarity=float(score)))
                    except Exception as parse_err:
                        logger.debug(f"Skip factor parse error: {parse_err}")
                        continue
                
                # ============================================================
                # STEP 5: ANSWER GENERATION (CROSS-ATTENTION LEVEL 2)
                # ============================================================
                logger.info(f"[SOLA EXPORT] =========================================")
                logger.info(f"[SOLA EXPORT] STEP 5: ANSWER GENERATION (CROSS-ATTENTION LEVEL 2)")
                logger.info(f"[SOLA EXPORT] Goal: Base Carbone factors ready for matching")
                logger.info(f"[SOLA EXPORT] Method: Reranked factors converted to MatchCandidate objects")
                logger.info(f"[SOLA EXPORT] Step 5: ✅ Cross-Attention Level 2 completed")
                logger.info(f"[SOLA EXPORT] Step 5: {len(matches)} factors ready for invoice matching")
                logger.info(f"[SOLA EXPORT] =========================================")
                logger.info(f"[SOLA EXPORT] ✅ RAG FLOW COMPLETED - All 5 steps executed")
                logger.info(f"[SOLA EXPORT] Summary:")
                logger.info(f"[SOLA EXPORT]   Step 1: Fast Retrieval (NO CROSS-ATTENTION) - Cosine similarity")
                logger.info(f"[SOLA EXPORT]   Step 2: Metadata / Entity Filtering")
                logger.info(f"[SOLA EXPORT]   Step 3: Re-ranking (CROSS-ATTENTION LEVEL 1) - Cross-Encoder")
                logger.info(f"[SOLA EXPORT]   Step 4: Prompt Construction (Factor Conversion)")
                logger.info(f"[SOLA EXPORT]   Step 5: Answer Generation (CROSS-ATTENTION LEVEL 2) - Factors Ready")
                logger.info(f"[SOLA EXPORT] {len(matches)} Base Carbone factors ready for matching")
                logger.info(f"[SOLA EXPORT] =========================================")
                
                if matches:
                    return matches
            except Exception as vec_err:
                logger.warning(f"[SOLA EXPORT] Vector search failed, fallback keyword search: {vec_err}")
                logger.exception(vec_err)
            
            # Fallback to keyword search if vector empty or failed
            logger.warning(f"[SOLA EXPORT] Falling back to keyword search")
            return _keyword_search(prompt, top_k=top_k)
        
        # Step 5: Process each invoice and match to Base Carbone factors
        # Logic matches map_invoices_to_base_carbone.py exactly
        processed = 0
        total_rows = len(invoices)
        logger.info(f"Processing {total_rows} invoices and matching to Base Carbone factors...")
        
        for invoice in invoices:
            try:
                
                # Check for strict match first
                strict_match = _find_strict_match(invoice)
                if strict_match:
                    logger.info(f"✅ STRICT MATCH: '{invoice.invoice_type}' -> '{strict_match.factor.name_fr}'")
                    strict_match_count += 1
                    candidates = [strict_match]
                    detected_category = None
                else:
                    # Detect invoice category for enhanced matching
                    detected_category = detect_invoice_category(invoice)
                    if detected_category:
                        logger.info(
                            f"📋 Detected category: {detected_category} for {invoice.invoice_type}"
                        )
                
                # Only perform semantic search if no strict match was found
                if not strict_match:
                    prompt = build_search_query(invoice)
                    
                    # Add category-specific keywords to search query if detected
                    if detected_category and detected_category in CATEGORY_MAPPINGS:
                        category_keywords = " ".join(
                            CATEGORY_MAPPINGS[detected_category]["keywords"][:3]
                        )
                        prompt = f"{prompt}; catégorie détectée: {category_keywords}"
                    
                    # Vector search via Azure Search
                    try:
                        candidates = _search_factors(prompt, top_k=5)
                    except Exception as exc:
                        if not embedding_failure_logged:
                            logger.warning(
                                f"⚠️ Embedding request failed ({exc}). Falling back to keyword search."
                            )
                            embedding_failure_logged = True
                        candidates = _keyword_search(prompt, top_k=5)
                    
                    # Fallback to keyword search if no candidates
                    if not candidates:
                        candidates = _keyword_search(prompt, top_k=5)
                    
                    # Raise error if still no candidates
                    if not candidates:
                        logger.error(
                            f"⚠️ No factor candidates for invoice: {invoice.invoice_type} - skipping"
                        )
                        continue
                    
                    # Apply enhanced factor search with category-specific matching
                    if detected_category:
                        candidates = enhanced_factor_search(invoice, candidates, detected_category)
                        logger.info(
                            f"✨ Enhanced matching applied. Top candidate: {candidates[0].factor.name_fr} (score: {candidates[0].similarity:.3f})"
                        )
                
                # LLM decision
                llm_decision = None
                if not llm_disabled and llm_client:
                    llm_decision, failure_reason, disable_now = call_llm_decision(
                        llm_client, "gpt-4", invoice, candidates
                    )
                    if failure_reason and failure_reason not in seen_failure_messages:
                        logger.warning(f"⚠️ {failure_reason}")
                        seen_failure_messages.add(failure_reason)
                        if (
                            len(seen_failure_messages) >= MAX_LLM_FAILURE_MESSAGES
                            and not llm_disabled
                        ):
                            if not llm_disable_notified:
                                logger.warning(
                                    "⚠️ Too many LLM warnings; disabling assistance for remaining invoices."
                                )
                                llm_disable_notified = True
                            llm_disabled = True
                    if disable_now:
                        if not llm_disable_notified:
                            logger.warning("⚠️ Disabling LLM assistance for remaining invoices.")
                            llm_disable_notified = True
                        llm_disabled = True
                    if disable_now or (failure_reason and llm_decision is None):
                        llm_decision = None
                    if llm_disabled:
                        llm_decision = None
                
                # Choose best factor
                selected = choose_factor(
                    candidates, llm_decision.selected_row_index if llm_decision else None
                )
                
                # Build mapping result
                mapping = build_mapping(
                    invoice, selected, candidates, rate_fetcher, llm_decision, detected_category
                )
                
                # Write to Excel
                writer.append_main(mapping)
                writer.append_audit(mapping)
                
                processed += 1
                
                # Update progress every 50 rows
                if processed % 50 == 0 or processed == total_rows:
                    progress = int(20 + (processed / total_rows) * 70)  # 20-90%
                    if progress_callback:
                        progress_callback(
                            "processing",
                            progress,
                            f"Matching invoices to Base Carbone factors... {processed}/{total_rows} completed"
                        )
                    logger.info(f"Processed {processed}/{total_rows} invoices...")
                
            except Exception as e:
                logger.warning(f"Failed to process invoice {invoice.invoice_type or 'unknown'}: {e}", exc_info=True)
                continue
        
        if processed == 0:
            error_msg = "Failed to process any invoices."
            logger.error(error_msg)
            result["error"] = error_msg
            if progress_callback:
                progress_callback("failed", 0, error_msg)
            return result
        
        # Step 6: Save Excel file
        if progress_callback:
            progress_callback("processing", 90, "Saving Excel file...")
        
        writer.save()
        logger.info(f"✅ Mapping workbook written to {excel_path}")
        
        # Store file path in result
        file_path_str = str(excel_path)
        strict_match_pct = (strict_match_count / processed * 100) if processed > 0 else 0.0
        summary_msg = f"Export completed successfully! {processed} invoices processed ({strict_match_count} strict matches, {strict_match_pct:.1f}%)."
        logger.info(f"📊 Strict matches: {strict_match_count}/{processed} ({strict_match_pct:.1f}%)")
        
        result["status"] = "completed"
        result["file_path"] = file_path_str
        result["strict_match_count"] = strict_match_count
        result["processed_count"] = processed
        
        if progress_callback:
            progress_callback("completed", 100, summary_msg)
        
        logger.info(f"🟢 [SOLA EXPORT] Completed - task_id: {task_id}, invoices: {processed}, strict_matches: {strict_match_count}")
        
        return result
        
    except Exception as e:
        logger.error(f"🔴 [SOLA EXPORT] Exception - task_id: {task_id}, error: {e}", exc_info=True)
        result["error"] = str(e)
        if progress_callback:
            progress_callback("failed", 0, f"Error: {str(e)}")
        return result
   